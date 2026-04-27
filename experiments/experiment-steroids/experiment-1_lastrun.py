#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
This experiment was created using PsychoPy3 Experiment Builder (v2026.1.3),
    on April 27, 2026, at 10:48
If you publish work using this script the most relevant publication is:

    Peirce J, Gray JR, Simpson S, MacAskill M, Höchenberger R, Sogo H, Kastman E, Lindeløv JK. (2019) 
        PsychoPy2: Experiments in behavior made easy Behav Res 51: 195. 
        https://doi.org/10.3758/s13428-018-01193-y

"""

# --- Import packages ---
from psychopy import locale_setup
from psychopy import prefs
from psychopy import plugins
plugins.activatePlugins()
from psychopy import sound, gui, visual, core, data, event, logging, clock, colors, layout, hardware
from psychopy.tools import environmenttools
from psychopy.constants import (
    NOT_STARTED, STARTED, PLAYING, PAUSED, STOPPED, STOPPING, FINISHED, PRESSED, 
    RELEASED, FOREVER, priority
)

import numpy as np  # whole numpy lib is available, prepend 'np.'
from numpy import (sin, cos, tan, log, log10, pi, average,
                   sqrt, std, deg2rad, rad2deg, linspace, asarray)
from numpy.random import random, randint, normal, shuffle, choice as randchoice
import os  # handy system and path functions
import sys  # to get file system encoding

from psychopy.hardware import keyboard
from psychopy_apparatus.hardware.apparatus import Apparatus
import os

# Run 'Before Experiment' code from setup_constants
# IMPORT MODULES ----
import random
import math
import numpy as np
import sounddevice as sd
import threading
import pandas as pd

# IMPORT SUBMODULES ----
# from scipy.stats import truncnorm, nur bei truncated normal distribution
from psychopy.colors import Color
from scipy.stats import norm
from itertools import product
from openpyxl import Workbook
from openpyxl.styles import PatternFill
# Run 'Before Experiment' code from setup_color_utils
def rgb_to_lab(rgb):
    """
    Convert color from RGB to CIELAB space. 
    
    Parameters
    ----------
    rgb : tuple or list
        A tuple/list of three values representing the RGB color.
        Accepts either normalized (0-1) or rgb255 (0-255) format.
        Auto-detects format: if any value > 1, assumes 0-255 range.
    """
    # Auto-normalize if input is in 0-255 range
    if any(c > 1 for c in rgb):
        rgb = tuple(c / 255 for c in rgb)
    
    def pivot_rgb(c):
        return ((c + 0.055)/1.055)**2.4 if c > 0.04045 else c/12.92
    
    R, G, B = [pivot_rgb(v) for v in rgb]
    
    X = 0.4124564*R + 0.3575761*G + 0.1804375*B
    Y = 0.2126729*R + 0.7151522*G + 0.0721750*B
    Z = 0.0193339*R + 0.1191920*G + 0.9503041*B
    
    X /= 0.95047
    Z /= 1.08883
    
    def f(t):
        return t**(1/3) if t > 0.008856 else (7.787*t + 16/116)
    
    L = 116 * f(Y) - 16
    a = 500 * (f(X) - f(Y))
    b = 200 * (f(Y) - f(Z))
    
    return (L, a, b)

def delta_e2000(lab1, lab2):
    """
    Calculate the Delta E 2000 color difference between two CIELAB colors.
    
    Parameters
    ----------
    lab1 : tuple
        A tuple of three floats representing the first CIELAB color (L, a, b).
    lab2 : tuple
        A tuple of three floats representing the second CIELAB color (L, a, b).
    """
    L1, a1, b1 = lab1
    L2, a2, b2 = lab2
    
    avg_L = (L1 + L2) / 2
    C1 = math.sqrt(a1*a1 + b1*b1)
    C2 = math.sqrt(a2*a2 + b2*b2)
    avg_C = (C1 + C2) / 2
    
    G = 0.5 * (1 - math.sqrt((avg_C**7)/((avg_C**7) + (25**7))))
    a1p = a1 * (1 + G)
    a2p = a2 * (1 + G)
    
    C1p = math.sqrt(a1p*a1p + b1*b1)
    C2p = math.sqrt(a2p*a2p + b2*b2)
    avg_Cp = (C1p + C2p) / 2
    
    h1p = (math.degrees(math.atan2(b1, a1p)) + 360) % 360
    h2p = (math.degrees(math.atan2(b2, a2p)) + 360) % 360
    
    dhp = h2p - h1p
    if abs(dhp) > 180:
        dhp -= 360 * math.copysign(1, dhp)
    dHp = 2 * math.sqrt(C1p*C2p) * math.sin(math.radians(dhp/2))
    
    avg_hp = (h1p + dhp/2) % 360
    
    T = 1 - 0.17*math.cos(math.radians(avg_hp - 30)) \
          + 0.24*math.cos(math.radians(2*avg_hp)) \
          + 0.32*math.cos(math.radians(3*avg_hp + 6)) \
          - 0.20*math.cos(math.radians(4*avg_hp - 63))
    
    Sl = 1 + (0.015*((avg_L - 50)**2)) / math.sqrt(20 + (avg_L - 50)**2)
    Sc = 1 + 0.045 * avg_Cp
    Sh = 1 + 0.015 * avg_Cp * T
    
    Rt = -2 * math.sqrt((avg_Cp**7)/((avg_Cp**7)+(25**7))) * \
         math.sin(math.radians(60 * math.exp(-(((avg_hp - 275)/25)**2))))
    
    dE = math.sqrt(
        (L2-L1)**2 / (Sl**2) +
        (C2p-C1p)**2 / (Sc**2) +
        (dHp)**2 / (Sh**2) +
        Rt * (C2p-C1p)/Sc * dHp/Sh
    )
    return dE

def generate_distractors(target_rgb, deltaE_mid, n):
    """
    Generate n distractor colors that are approximately deltaE_mid away from target.
    
    Parameters
    ----------
    target_rgb : tuple/list
        RGB color in 0-1 (normalized) or 0-255 format (auto-detected).
    deltaE_mid : float
        Target Delta E 2000 distance from target color.
    n : int
        Number of distractor colors to generate.
    
    Returns
    -------
    list of lists
        Distractor colors in rgb255 format (0-255 ints).
    """
    target_lab = rgb_to_lab(target_rgb)
    distractors = []
    attempts = 0
    
    while len(distractors) < n and attempts < 50000:
        attempts += 1
        rgb = (random.random(), random.random(), random.random())
        lab = rgb_to_lab(rgb)
        dE = delta_e2000(target_lab, lab)
        
        if abs(dE - deltaE_mid) <= 1.5:
            distractors.append([int(c * 255) for c in rgb])  # Convert to rgb255
    
    if len(distractors) < n:
        print("WARNUNG: nicht genug Farben im DeltaE-Band gefunden.")
        while len(distractors) < n:
            distractors.append([int(c * 255) for c in (random.random(), random.random(), random.random())])
    
    return distractors
# --- Setup global variables (available in all functions) ---
# create a device manager to handle hardware (keyboards, mice, mirophones, speakers, etc.)
deviceManager = hardware.DeviceManager()
# ensure that relative paths start from the same directory as this script
_thisDir = os.path.dirname(os.path.abspath(__file__))
# store info about the experiment session
psychopyVersion = '2026.1.3'
expName = 'maxForce'  # from the Builder filename that created this script
expVersion = ''
# a list of functions to run when the experiment ends (starts off blank)
runAtExit = []
# information about this experiment
expInfo = {
    'participant': f"{randint(0, 999999):06.0f}",
    'session': '001',
    'date|hid': data.getDateStr(),
    'expName|hid': expName,
    'expVersion|hid': expVersion,
    'psychopyVersion|hid': psychopyVersion,
}

# --- Define some variables which will change depending on pilot mode ---
'''
To run in pilot mode, either use the run/pilot toggle in Builder, Coder and Runner, 
or run the experiment with `--pilot` as an argument. To change what pilot 
#mode does, check out the 'Pilot mode' tab in preferences.
'''
# work out from system args whether we are running in pilot mode
PILOTING = core.setPilotModeFromArgs()
# start off with values from experiment settings
_fullScr = True
_winSize = [1280, 800]
# if in pilot mode, apply overrides according to preferences
if PILOTING:
    # force windowed mode
    if prefs.piloting['forceWindowed']:
        _fullScr = False
        # set window size
        _winSize = prefs.piloting['forcedWindowSize']
    # replace default participant ID
    if prefs.piloting['replaceParticipantID']:
        expInfo['participant'] = 'pilot'

def showExpInfoDlg(expInfo):
    """
    Show participant info dialog.
    Parameters
    ==========
    expInfo : dict
        Information about this experiment.
    
    Returns
    ==========
    dict
        Information about this experiment.
    """
    # show participant info dialog
    dlg = gui.DlgFromDict(
        dictionary=expInfo, sortKeys=False, title=expName, alwaysOnTop=True
    )
    if dlg.OK == False:
        core.quit()  # user pressed cancel
    # return expInfo
    return expInfo


def setupData(expInfo, dataDir=None):
    """
    Make an ExperimentHandler to handle trials and saving.
    
    Parameters
    ==========
    expInfo : dict
        Information about this experiment, created by the `setupExpInfo` function.
    dataDir : Path, str or None
        Folder to save the data to, leave as None to create a folder in the current directory.    
    Returns
    ==========
    psychopy.data.ExperimentHandler
        Handler object for this experiment, contains the data to save and information about 
        where to save it to.
    """
    # remove dialog-specific syntax from expInfo
    for key, val in expInfo.copy().items():
        newKey, _ = data.utils.parsePipeSyntax(key)
        expInfo[newKey] = expInfo.pop(key)
    
    # data file name stem = absolute path + name; later add .psyexp, .csv, .log, etc
    if dataDir is None:
        dataDir = _thisDir
    filename = u'data/%s_%s_%s' % (expInfo['participant'], expName, expInfo['date'])
    # make sure filename is relative to dataDir
    if os.path.isabs(filename):
        dataDir = os.path.commonprefix([dataDir, filename])
        filename = os.path.relpath(filename, dataDir)
    
    # an ExperimentHandler isn't essential but helps with data saving
    thisExp = data.ExperimentHandler(
        name=expName, version=expVersion,
        extraInfo=expInfo, runtimeInfo=None,
        originPath='C:\\Users\\Maik Bieleke\\My Drive\\Labor\\temp\\experiments\\experiment-1\\experiment-1_lastrun.py',
        savePickle=True, saveWideText=True,
        dataFileName=dataDir + os.sep + filename, sortColumns='time'
    )
    # store pilot mode in data file
    thisExp.addData('piloting', PILOTING, priority=priority.LOW)
    thisExp.setPriority('thisRow.t', priority.CRITICAL)
    thisExp.setPriority('expName', priority.LOW)
    # return experiment handler
    return thisExp


def setupLogging(filename):
    """
    Setup a log file and tell it what level to log at.
    
    Parameters
    ==========
    filename : str or pathlib.Path
        Filename to save log file and data files as, doesn't need an extension.
    
    Returns
    ==========
    psychopy.logging.LogFile
        Text stream to receive inputs from the logging system.
    """
    # set how much information should be printed to the console / app
    if PILOTING:
        logging.console.setLevel(
            prefs.piloting['pilotConsoleLoggingLevel']
        )
    else:
        logging.console.setLevel('warning')
    # save a log file for detail verbose info
    logFile = logging.LogFile(filename+'.log')
    if PILOTING:
        logFile.setLevel(
            prefs.piloting['pilotLoggingLevel']
        )
    else:
        logFile.setLevel(
            logging.getLevel('debug')
        )
    
    return logFile


def setupWindow(expInfo=None, win=None):
    """
    Setup the Window
    
    Parameters
    ==========
    expInfo : dict
        Information about this experiment, created by the `setupExpInfo` function.
    win : psychopy.visual.Window
        Window to setup - leave as None to create a new window.
    
    Returns
    ==========
    psychopy.visual.Window
        Window in which to run this experiment.
    """
    if PILOTING:
        logging.debug('Fullscreen settings ignored as running in pilot mode.')
    
    if win is None:
        # if not given a window to setup, make one
        win = visual.Window(
            size=_winSize, fullscr=_fullScr, screen=1,
            winType='pyglet', allowGUI=True, allowStencil=False,
            monitor='testMonitor', color=(-1.0000, -1.0000, -1.0000), colorSpace='rgb',
            backgroundImage='', backgroundFit='none',
            blendMode='avg', useFBO=True,
            units='height',
            checkTiming=False  # we're going to do this ourselves in a moment
        )
    else:
        # if we have a window, just set the attributes which are safe to set
        win.color = (-1.0000, -1.0000, -1.0000)
        win.colorSpace = 'rgb'
        win.backgroundImage = ''
        win.backgroundFit = 'none'
        win.units = 'height'
    if expInfo is not None:
        # get/measure frame rate if not already in expInfo
        if win._monitorFrameRate is None:
            win._monitorFrameRate = win.getActualFrameRate(infoMsg='Experiment wird vorbereitet...')
        expInfo['frameRate'] = win._monitorFrameRate
    win.hideMessage()
    if PILOTING:
        # show a visual indicator if we're in piloting mode
        if prefs.piloting['showPilotingIndicator']:
            win.showPilotingIndicator()
        # always show the mouse in piloting mode
        if prefs.piloting['forceMouseVisible']:
            win.mouseVisible = True
    
    return win


def setupDevices(expInfo, thisExp, win):
    """
    Setup whatever devices are available (mouse, keyboard, speaker, eyetracker, etc.) and add them to 
    the device manager (deviceManager)
    
    Parameters
    ==========
    expInfo : dict
        Information about this experiment, created by the `setupExpInfo` function.
    thisExp : psychopy.data.ExperimentHandler
        Handler object for this experiment, contains the data to save and information about 
        where to save it to.
    win : psychopy.visual.Window
        Window in which to run this experiment.
    Returns
    ==========
    bool
        True if completed successfully.
    """
    # --- Setup input devices ---
    ioConfig = {}
    ioSession = ioServer = eyetracker = None
    
    # store ioServer object in the device manager
    deviceManager.ioServer = ioServer
    
    # create a default keyboard (e.g. to check for escape)
    if deviceManager.getDevice('defaultKeyboard') is None:
        deviceManager.addDevice(
            deviceClass='keyboard', deviceName='defaultKeyboard', backend='ptb'
        )
    # initialize 'apparatus'
    deviceManager.addDevice(
        deviceName='apparatus',
        deviceClass='psychopy.hardware.ApparatusDevice',
        port='COM6',
        baudrate='921600',
        simulate=False,
        debug=False
    )
    # return True if completed successfully
    return True

def pauseExperiment(thisExp, win=None, timers=[], currentRoutine=None):
    """
    Pause this experiment, preventing the flow from advancing to the next routine until resumed.
    
    Parameters
    ==========
    thisExp : psychopy.data.ExperimentHandler
        Handler object for this experiment, contains the data to save and information about 
        where to save it to.
    win : psychopy.visual.Window
        Window for this experiment.
    timers : list, tuple
        List of timers to reset once pausing is finished.
    currentRoutine : psychopy.data.Routine
        Current Routine we are in at time of pausing, if any. This object tells PsychoPy what Components to pause/play/dispatch.
    """
    # if we are not paused, do nothing
    if thisExp.status != PAUSED:
        return
    
    # start a timer to figure out how long we're paused for
    pauseTimer = core.Clock()
    # pause any playback components
    if currentRoutine is not None:
        for comp in currentRoutine.getPlaybackComponents():
            comp.pause()
    # make sure we have a keyboard
    defaultKeyboard = deviceManager.getDevice('defaultKeyboard')
    if defaultKeyboard is None:
        defaultKeyboard = deviceManager.addKeyboard(
            deviceClass='keyboard',
            deviceName='defaultKeyboard',
            backend='PsychToolbox',
        )
    # run a while loop while we wait to unpause
    while thisExp.status == PAUSED:
        # check for quit (typically the Esc key)
        if defaultKeyboard.getKeys(keyList=['escape']):
            endExperiment(thisExp, win=win)
        # dispatch messages on response components
        if currentRoutine is not None:
            for comp in currentRoutine.getDispatchComponents():
                comp.device.dispatchMessages()
        # sleep 1ms so other threads can execute
        clock.time.sleep(0.001)
    # if stop was requested while paused, quit
    if thisExp.status == FINISHED:
        endExperiment(thisExp, win=win)
    # resume any playback components
    if currentRoutine is not None:
        for comp in currentRoutine.getPlaybackComponents():
            comp.play()
    # reset any timers
    for timer in timers:
        timer.addTime(-pauseTimer.getTime())


def run(expInfo, thisExp, win, globalClock=None, thisSession=None):
    """
    Run the experiment flow.
    
    Parameters
    ==========
    expInfo : dict
        Information about this experiment, created by the `setupExpInfo` function.
    thisExp : psychopy.data.ExperimentHandler
        Handler object for this experiment, contains the data to save and information about 
        where to save it to.
    psychopy.visual.Window
        Window in which to run this experiment.
    globalClock : psychopy.core.clock.Clock or None
        Clock to get global time from - supply None to make a new one.
    thisSession : psychopy.session.Session or None
        Handle of the Session object this experiment is being run from, if any.
    """
    # mark experiment as started
    thisExp.status = STARTED
    # update experiment info
    expInfo['date'] = data.getDateStr()
    expInfo['expName'] = expName
    expInfo['expVersion'] = expVersion
    expInfo['psychopyVersion'] = psychopyVersion
    # make sure window is set to foreground to prevent losing focus
    win.winHandle.activate()
    # make sure variables created by exec are available globally
    exec = environmenttools.setExecEnvironment(globals())
    # get device handles from dict of input devices
    ioServer = deviceManager.ioServer
    # get/create a default keyboard (e.g. to check for escape)
    defaultKeyboard = deviceManager.getDevice('defaultKeyboard')
    if defaultKeyboard is None:
        deviceManager.addDevice(
            deviceClass='keyboard', deviceName='defaultKeyboard', backend='PsychToolbox'
        )
    eyetracker = deviceManager.getDevice('eyetracker')
    # make sure we're running in the directory for this experiment
    os.chdir(_thisDir)
    # get filename from ExperimentHandler for convenience
    filename = thisExp.dataFileName
    frameTolerance = 0.001  # how close to onset before 'same' frame
    endExpNow = False  # flag for 'escape' or other condition => quit the exp
    # get frame duration from frame rate in expInfo
    if 'frameRate' in expInfo and expInfo['frameRate'] is not None:
        frameDur = 1.0 / round(expInfo['frameRate'])
    else:
        frameDur = 1.0 / 60.0  # could not measure, so guess
    
    # Start Code - component code to be run after the window creation
    
    # --- Initialize components for Routine "rr_setup" ---
    # Run 'Begin Experiment' code from setup_constants
    # DEFINITIONS AND CONSTANTS ----
    NUM_MVC      = 3  # How many MVC trials to conduct
    NUM_DOMAINS  = 0 #1  # How often each domain is repeated
    NUM_BLOCKS   = 3  # How often each block within the domain is repeated
    NUM_LEVELS   = 6  # How often each level within the block is repeated
    NUM_PALETTES = 24 # How many palettes are generated per trial
    NUM_TARGETS  = 4  # How many targets appear in NUM_PALETTES palettes
    NUM_DISTRACTORS = 8
    
    # MIN, MAX, AND RANGE OF COGNITIVE AND PHYSICAL DEMAND
    PHYSICAL_DEMAND_MIN     = 0.03
    PHYSICAL_DEMAND_MAX     = 0.40
    COGNITIVE_DEMAND_MIN    = 10
    COGNITIVE_DEMAND_MAX    = 60
    PHYSICAL_DEMAND_LEVELS  = np.round(np.linspace(PHYSICAL_DEMAND_MIN, PHYSICAL_DEMAND_MAX, 4), 2).tolist()
    COGNITIVE_DEMAND_LEVELS = np.round(np.linspace(COGNITIVE_DEMAND_MIN, COGNITIVE_DEMAND_MAX, 4), 2).tolist()
    
    TOLERANCE = 15.0
    
    # TARGET COLORS ----
    TARGET_COLORS_RGB255 = [[255, 0, 0], [0, 255, 0], [0, 0, 255]]
    COLOR_NAMES = {
        (255, 0, 0):   "Red",
        (0, 255, 0):   "Green",
        (0, 0, 255):   "Blue",
    }
    
    INNER_HOLES = list(range(8))
    OUTER_HOLES = list(range(9, 21))
    
    # TIMING CONSTANTS ----
    PALETTE_DURATION_SEC = 0.416
    COUNTDOWN_STEP_SEC = 0.5
    
    # AUDIO CONSTANTS ----
    FS          = 48000
    CHANNELS    = 1
    FREQ_LOW    = 440.0
    FREQ_HIGH   = 880.0
    AMP_MAX     = 1.0
    AMP_SMOOTH  = 0.1
    STATE_HOLD  = 0.1
    # Run 'Begin Experiment' code from setup_audio
    global _state_lock, _target_freq, _target_amp, _cur_amp, _phase, _stream
    
    # Audio parameters
    FS          = 48000
    CHANNELS    = 1
    FREQ_LOW    = 440.0
    FREQ_HIGH   = 880.0
    AMP_MAX     = 1.0
    AMP_SMOOTH  = 0.1
    STATE_HOLD  = 0.1
    
    # Shared state
    _state_lock  = threading.Lock()
    _target_freq = FREQ_LOW
    _target_amp  = 0.0
    _cur_amp     = 0.0
    _phase       = 0.0
    
    def audio_callback(outdata, frames, time, status):
        global _phase, _cur_amp, _target_amp, _target_freq
        with _state_lock:
            tf = float(_target_freq)
            ta = float(_target_amp)
        _cur_amp  = (1.0 - AMP_SMOOTH) * _cur_amp + AMP_SMOOTH * ta
        t         = (np.arange(frames) / FS).astype(np.float32)
        phase_inc = 2.0 * np.pi * tf
        y         = np.sin(_phase + phase_inc * t).astype(np.float32)
        _phase    = float((_phase + phase_inc * (frames / FS)) % (2.0 * np.pi))
        outdata[:] = (_cur_amp * AMP_MAX * y).reshape(-1, 1)
    
    _stream = sd.OutputStream(
        samplerate = FS,
        channels   = CHANNELS,
        dtype      = "float32",
        callback   = audio_callback,
    )
    if not _stream.active:
        _stream.start()
    # Run 'Begin Experiment' code from setup_random
    # =============================================================================
    # PALETTE GENERATION
    # =============================================================================
    def generate_trial_palettes(target_color_rgb255, cognitive_demand_level,
                                 n_palettes = NUM_PALETTES, n_targets = NUM_TARGETS,
                                 n_holes = NUM_DISTRACTORS):
        """
        Generate all palettes for one trial.
        - n_targets palettes contain the target color at a random hole position
        - remaining palettes contain only distractors
        """
        target_palette_indices = set(random.sample(range(n_palettes), n_targets))
    
        palettes = []
        for i in range(n_palettes):
            if i in target_palette_indices:
                distractors = generate_distractors(target_color_rgb255, cognitive_demand_level, n=n_holes - 1)
                target_pos  = random.randint(0, n_holes - 1)
                palette     = distractors.copy()
                palette.insert(target_pos, target_color_rgb255)
            else:
                palette = generate_distractors(target_color_rgb255, cognitive_demand_level, n=n_holes)
            palettes.append(palette)
    
        return palettes
    
    # =============================================================================
    # CREATE TRIALS
    # =============================================================================
    blocks = {b: [] for b in range(NUM_BLOCKS)}
    for physical_demand_level, cognitive_demand_level in product(PHYSICAL_DEMAND_LEVELS, COGNITIVE_DEMAND_LEVELS):
        for block, target_color_rgb255 in zip(blocks, random.sample(TARGET_COLORS_RGB255, 3)):
            blocks[block].append({
                "physical_demand_level":  physical_demand_level,
                "cognitive_demand_level": cognitive_demand_level,
                "target_color_rgb255":    target_color_rgb255,
            })
    
    for block in blocks.values():
        random.shuffle(block)
    
    # =============================================================================
    # CREATE PALETTES
    # =============================================================================
    for b, trials in blocks.items():
        for trial in trials:
            palettes = generate_trial_palettes(
                target_color_rgb255    = trial["target_color_rgb255"],
                cognitive_demand_level = trial["cognitive_demand_level"],
            )
            target = trial["target_color_rgb255"]
            target_palettes = [i for i, p in enumerate(palettes) if target in p]
            target_holes    = [p.index(target) for i, p in enumerate(palettes) if target in p]
            
            trial["palettes"]            = palettes
            trial["palette_has_target"]  = [target in p for p in palettes]
            trial["palette_target_hole"] = [p.index(target) if target in p else None for p in palettes]
    
    # =============================================================================
    # BALANCE CHECK
    # =============================================================================
    all_trials = []
    for b, trials in blocks.items():
        for t in trials:
            all_trials.append({**{k: v for k, v in t.items() if k != "palettes"}, "block": b})
    df_check = pd.DataFrame(all_trials)
    df_check["color"] = df_check["target_color_rgb255"].apply(tuple)
    
    lines = ["=" * 50, "TRIAL BALANCE CHECK", "=" * 50]
    for factor in ["physical_demand_level", "cognitive_demand_level", "color"]:
        lines.append(f"\n--- {factor} ---")
        table = df_check.groupby(["block", factor]).size().unstack(fill_value=0)
        lines.append(table.to_string())
        lines.append(f"Totals: {df_check[factor].value_counts().sort_index().to_dict()}")
    lines.append("\nTotal trials per block:")
    lines.append(df_check.groupby("block").size().to_string())
    print("\n".join(lines))
    
    # =============================================================================
    # BUILD LONG FORMAT ROWS
    # =============================================================================
    rows = []
    for b, trials in blocks.items():
        for trial_idx, trial in enumerate(trials):
            for palette_idx, palette in enumerate(trial["palettes"]):
                has_target = trial["target_color_rgb255"] in palette
                for hole_idx, color in enumerate(palette):
                    rows.append({
                        "block":                  b,
                        "trial_idx":              trial_idx,
                        "physical_demand_level":  trial["physical_demand_level"],
                        "cognitive_demand_level": trial["cognitive_demand_level"],
                        "target_color_rgb255":    trial["target_color_rgb255"],
                        "palette_idx":            palette_idx,
                        "is_target_palette":      has_target,
                        "hole_idx":               hole_idx,
                        "hole_color_rgb255":      color,
                        "is_target_hole":         color == trial["target_color_rgb255"],
                        # Placeholders for data
                        "is_selected":              None,
                        "selected_rt":              None,
                        "is_hit":                   None,
                        "is_false_alarm":           None,
                        "is_miss":                  None,
                        "is_correct_rejection":     None,
                        "hits":                     None,
                        "false_alarms":             None,
                        "misses":                   None,
                        "correct_rejections":       None,
                        "d_prime":                  None
                    })
    
    # Create an index for fast lookup
    row_index = {}
    palette_row_index = {}
    for i, row in enumerate(rows):
        hole_key = (row["block"], row["trial_idx"], row["palette_idx"], row["hole_idx"])
        row_index[hole_key] = i
        palette_key = (row["block"], row["trial_idx"], row["palette_idx"])
        if palette_key not in palette_row_index:
            palette_row_index[palette_key] = []
        palette_row_index[palette_key].append(i)
    
    # =============================================================================
    # EXPORT TO EXCEL WITH COLOR PREVIEW
    # =============================================================================
    # wb = Workbook()
    # ws = wb.active
    
    # headers = list(rows[0].keys()) + ["color_preview"]
    # ws.append(headers)
    
    # for row in rows:
    #     converted = {k: str(v) if isinstance(v, list) else v for k, v in row.items()}
    #     ws.append(list(converted.values()) + [""])
    # 
    #     color = row["hole_color_rgb255"]
    #     hex_color = "{:02X}{:02X}{:02X}".format(color[0], color[1], color[2])
    #     ws.cell(row=ws.max_row, column=len(headers)).fill = PatternFill(
    #         start_color=hex_color, end_color=hex_color, fill_type="solid"
    #     )
    # 
    # wb.save(thisExp.dataFileName + '_design.xlsx')
    
    pd.DataFrame(rows).to_csv(thisExp.dataFileName + '_design.csv', index=False)
    # print(f"\nSaved {len(rows)} rows to _design.csv")
    # print(f"Expected: {NUM_BLOCKS} blocks × 16 trials × {NUM_PALETTES} palettes × {NUM_DISTRACTORS} holes = {NUM_BLOCKS * 16 * NUM_PALETTES * NUM_DISTRACTORS} rows")
    # Run 'Begin Experiment' code from setup_ring
    # Main ring parameters
    ring_radius = 0.075  # distance of small circles from center
    hole_radius = 0.02   # size of each small circle
    
    # Create the 8 hole stimuli
    hole_stimuli = []
    ring_center = (0.5, 0.0)
    for i in range(NUM_DISTRACTORS):
        angle = np.pi - 2 * np.pi * i / NUM_DISTRACTORS
        x = ring_center[0] + ring_radius * np.cos(angle)
        y = ring_center[1] + ring_radius * np.sin(angle)
        circle = visual.Circle(
            win,
            radius=hole_radius,
            pos=(x, y),
            fillColor=[128, 128, 128],
            colorSpace='rgb255',
            lineColor=None,
        )
        hole_stimuli.append(circle)
    
    # Create the labels
    hole_labels = []
    for i in range(NUM_DISTRACTORS):
        angle = np.pi - 2 * np.pi * i / NUM_DISTRACTORS
        x = ring_center[0] + ring_radius * np.cos(angle)
        y = ring_center[1] + ring_radius * np.sin(angle)
        label = visual.TextStim(
            win,
            text=str(i),
            pos=(x, y),
            height=0.02,
            color='white',
            bold=True,
        )
        hole_labels.append(label)
    
    # --- Initialize components for Routine "rr_start" ---
    start_text = visual.TextStim(win=win, name='start_text',
        text='Fertig - es kann losgehen!',
        font='Arial',
        pos=(0, 0), draggable=False, height=0.05, wrapWidth=None, ori=0.0, 
        color='white', colorSpace='rgb', opacity=None, 
        languageStyle='LTR',
        depth=0.0);
    start_button = visual.ButtonStim(win, 
        text='START', font='Arvo',
        pos=(0, -0.2),
        letterHeight=0.05,
        size=(0.25, 0.1), 
        ori=0.0
        ,borderWidth=0.0,
        fillColor=(-1.0000, -0.2157, -1.0000), borderColor=None,
        color='white', colorSpace='rgb',
        opacity=None,
        bold=True, italic=False,
        padding=None,
        anchor='center',
        name='start_button',
        depth=-1
    )
    start_button.buttonClock = core.Clock()
    
    # --- Initialize components for Routine "rr_countdown" ---
    
    countdown_lights = Apparatus('apparatus')
    countdown_display = visual.TextStim(win=win, name='countdown_display',
        text='',
        font='Arial',
        pos=(0, 0), draggable=False, height=0.15, wrapWidth=None, ori=0.0, 
        color='white', colorSpace='rgb', opacity=None, 
        languageStyle='LTR',
        depth=-2.0);
    
    # --- Initialize components for Routine "rr_maxforce" ---
    # Run 'Begin Experiment' code from maxforce_code
    # Initialize value for maximum voluntary contraction (MVC)
    mvc = 0
    
    maxforce_lights = Apparatus('apparatus')
    
    maxforce_force = Apparatus('apparatus')
    maxforce_display = visual.TextStim(win=win, name='maxforce_display',
        text='',
        font='Courier New',
        pos=(0, 0), draggable=False, height=0.04, wrapWidth=None, ori=0.0, 
        color='white', colorSpace='rgb', opacity=None, 
        languageStyle='LTR',
        depth=-3.0);
    
    # --- Initialize components for Routine "rr_mvc" ---
    mvc_display = visual.TextStim(win=win, name='mvc_display',
        text='',
        font='Courier New',
        pos=(0, 0), draggable=False, height=0.04, wrapWidth=None, ori=0.0, 
        color=(-1.0000, -0.2157, -1.0000), colorSpace='rgb', opacity=None, 
        languageStyle='LTR',
        depth=-1.0);
    mvc_continue = visual.ButtonStim(win, 
        text='Weiter', font='Arvo',
        pos=(0, -0.3),
        letterHeight=0.05,
        size=(0.25, 0.1), 
        ori=0.0
        ,borderWidth=0.0,
        fillColor=(-1.0000, -0.2157, -1.0000), borderColor=None,
        color='white', colorSpace='rgb',
        opacity=None,
        bold=True, italic=False,
        padding=None,
        anchor='center',
        name='mvc_continue',
        depth=-2
    )
    mvc_continue.buttonClock = core.Clock()
    
    # --- Initialize components for Routine "rr_countdown" ---
    
    countdown_lights = Apparatus('apparatus')
    countdown_display = visual.TextStim(win=win, name='countdown_display',
        text='',
        font='Arial',
        pos=(0, 0), draggable=False, height=0.15, wrapWidth=None, ori=0.0, 
        color='white', colorSpace='rgb', opacity=None, 
        languageStyle='LTR',
        depth=-2.0);
    
    # --- Initialize components for Routine "rr_target" ---
    
    target_lights = Apparatus(None)
    
    target_force = Apparatus(None)
    target_display = visual.TextStim(win=win, name='target_display',
        text='',
        font='Courier New',
        pos=(0, 0), draggable=False, height=0.04, wrapWidth=None, ori=0.0, 
        color='white', colorSpace='rgb', opacity=None, 
        languageStyle='LTR',
        depth=-3.0);
    
    # --- Initialize components for Routine "rr_trial" ---
    
    trial_reed = Apparatus('apparatus')
    
    trial_lights = Apparatus(None)
    
    trial_force = Apparatus(None)
    trial_display = visual.TextStim(win=win, name='trial_display',
        text='',
        font='Courier New',
        pos=(0, 0), draggable=False, height=0.04, wrapWidth=None, ori=0.0, 
        color='white', colorSpace='rgb', opacity=None, 
        languageStyle='LTR',
        depth=-4.0);
    
    # --- Initialize components for Routine "rr_dprime" ---
    feedback_keyboard = keyboard.Keyboard(deviceName='defaultKeyboard')
    feedback_display = visual.TextStim(win=win, name='feedback_display',
        text='',
        font='Courier New',
        pos=(0, 0), draggable=False, height=0.04, wrapWidth=None, ori=0.0, 
        color='white', colorSpace='rgb', opacity=None, 
        languageStyle='LTR',
        depth=-2.0);
    
    # --- Initialize components for Routine "rr_rating" ---
    rating_slider = visual.Slider(win=win, name='rating_slider',
        startValue=None, size=(1.0, 0.1), pos=(0, -0.2), units=win.units,
        labels=(0,1,2,3,4,5,6,7,8,9,10), ticks=(0,1,2,3,4,5,6,7,8,9,10), granularity=1.0,
        style='rating', styleTweaks=[], opacity=None,
        labelColor='LightGray', markerColor='Red', lineColor='White', colorSpace='rgb',
        font='Noto Sans', labelHeight=0.05,
        flip=False, ori=0.0, depth=0, readOnly=False)
    rating_keyboard = keyboard.Keyboard(deviceName='defaultKeyboard')
    rating_display = visual.TextStim(win=win, name='rating_display',
        text='',
        font='Arial',
        pos=(0, 0), draggable=False, height=0.05, wrapWidth=None, ori=0.0, 
        color='white', colorSpace='rgb', opacity=None, 
        languageStyle='LTR',
        depth=-2.0);
    
    # create some handy timers
    
    # global clock to track the time since experiment started
    if globalClock is None:
        # create a clock if not given one
        globalClock = core.Clock()
    if isinstance(globalClock, str):
        # if given a string, make a clock accoridng to it
        if globalClock == 'float':
            # get timestamps as a simple value
            globalClock = core.Clock(format='float')
        elif globalClock == 'iso':
            # get timestamps in ISO format
            globalClock = core.Clock(format='%Y-%m-%d_%H:%M:%S.%f%z')
        else:
            # get timestamps in a custom format
            globalClock = core.Clock(format=globalClock)
    if ioServer is not None:
        ioServer.syncClock(globalClock)
    logging.setDefaultClock(globalClock)
    if eyetracker is not None:
        eyetracker.enableEventReporting()
    # routine timer to track time remaining of each (possibly non-slip) routine
    routineTimer = core.Clock()
    win.flip()  # flip window to reset last flip timer
    # store the exact time the global clock started
    expInfo['expStart'] = data.getDateStr(
        format='%Y-%m-%d %Hh%M.%S.%f %z', fractionalSecondDigits=6
    )
    
    # --- Prepare to start Routine "rr_setup" ---
    # create an object to store info about Routine rr_setup
    rr_setup = data.Routine(
        name='rr_setup',
        components=[],
    )
    rr_setup.status = NOT_STARTED
    continueRoutine = True
    # update component parameters for each repeat
    # store start times for rr_setup
    rr_setup.tStartRefresh = win.getFutureFlipTime(clock=globalClock)
    rr_setup.tStart = globalClock.getTime(format='float')
    rr_setup.status = STARTED
    thisExp.addData('rr_setup.started', rr_setup.tStart)
    rr_setup.maxDuration = None
    # keep track of which components have finished
    rr_setupComponents = rr_setup.components
    for thisComponent in rr_setup.components:
        thisComponent.tStart = None
        thisComponent.tStop = None
        thisComponent.tStartRefresh = None
        thisComponent.tStopRefresh = None
        if hasattr(thisComponent, 'status'):
            thisComponent.status = NOT_STARTED
    # reset timers
    t = 0
    _timeToFirstFrame = win.getFutureFlipTime(clock="now")
    frameN = -1
    
    # --- Run Routine "rr_setup" ---
    thisExp.currentRoutine = rr_setup
    rr_setup.forceEnded = routineForceEnded = not continueRoutine
    while continueRoutine:
        # get current time
        t = routineTimer.getTime()
        tThisFlip = win.getFutureFlipTime(clock=routineTimer)
        tThisFlipGlobal = win.getFutureFlipTime(clock=None)
        frameN = frameN + 1  # number of completed frames (so 0 is the first frame)
        # update/draw components on each frame
        
        # check for quit (typically the Esc key)
        if defaultKeyboard.getKeys(keyList=["escape"]):
            thisExp.status = FINISHED
        if thisExp.status == FINISHED or endExpNow:
            endExperiment(thisExp, win=win)
            return
        # pause experiment here if requested
        if thisExp.status == PAUSED:
            pauseExperiment(
                thisExp=thisExp, 
                win=win, 
                timers=[routineTimer, globalClock], 
                currentRoutine=rr_setup,
            )
            # skip the frame we paused on
            continue
        
        # has a Component requested the Routine to end?
        if not continueRoutine:
            rr_setup.forceEnded = routineForceEnded = True
        # has the Routine been forcibly ended?
        if rr_setup.forceEnded or routineForceEnded:
            break
        # has every Component finished?
        continueRoutine = False
        for thisComponent in rr_setup.components:
            if hasattr(thisComponent, "status") and thisComponent.status != FINISHED:
                continueRoutine = True
                break  # at least one component has not yet finished
        
        # refresh the screen
        if continueRoutine:  # don't flip if this routine is over or we'll get a blank screen
            win.flip()
    
    # --- Ending Routine "rr_setup" ---
    for thisComponent in rr_setup.components:
        if hasattr(thisComponent, "setAutoDraw"):
            thisComponent.setAutoDraw(False)
    # store stop times for rr_setup
    rr_setup.tStop = globalClock.getTime(format='float')
    rr_setup.tStopRefresh = tThisFlipGlobal
    thisExp.addData('rr_setup.stopped', rr_setup.tStop)
    thisExp.nextEntry()
    # the Routine "rr_setup" was not non-slip safe, so reset the non-slip timer
    routineTimer.reset()
    
    # --- Prepare to start Routine "rr_start" ---
    # create an object to store info about Routine rr_start
    rr_start = data.Routine(
        name='rr_start',
        components=[start_text, start_button],
    )
    rr_start.status = NOT_STARTED
    continueRoutine = True
    # update component parameters for each repeat
    # reset start_button to account for continued clicks & clear times on/off
    start_button.reset()
    # store start times for rr_start
    rr_start.tStartRefresh = win.getFutureFlipTime(clock=globalClock)
    rr_start.tStart = globalClock.getTime(format='float')
    rr_start.status = STARTED
    thisExp.addData('rr_start.started', rr_start.tStart)
    rr_start.maxDuration = None
    # keep track of which components have finished
    rr_startComponents = rr_start.components
    for thisComponent in rr_start.components:
        thisComponent.tStart = None
        thisComponent.tStop = None
        thisComponent.tStartRefresh = None
        thisComponent.tStopRefresh = None
        if hasattr(thisComponent, 'status'):
            thisComponent.status = NOT_STARTED
    # reset timers
    t = 0
    _timeToFirstFrame = win.getFutureFlipTime(clock="now")
    frameN = -1
    
    # --- Run Routine "rr_start" ---
    thisExp.currentRoutine = rr_start
    rr_start.forceEnded = routineForceEnded = not continueRoutine
    while continueRoutine:
        # get current time
        t = routineTimer.getTime()
        tThisFlip = win.getFutureFlipTime(clock=routineTimer)
        tThisFlipGlobal = win.getFutureFlipTime(clock=None)
        frameN = frameN + 1  # number of completed frames (so 0 is the first frame)
        # update/draw components on each frame
        
        # *start_text* updates
        
        # if start_text is starting this frame...
        if start_text.status == NOT_STARTED and tThisFlip >= 0.0-frameTolerance:
            # keep track of start time/frame for later
            start_text.frameNStart = frameN  # exact frame index
            start_text.tStart = t  # local t and not account for scr refresh
            start_text.tStartRefresh = tThisFlipGlobal  # on global time
            win.timeOnFlip(start_text, 'tStartRefresh')  # time at next scr refresh
            # add timestamp to datafile
            thisExp.timestampOnFlip(win, 'start_text.started')
            # update status
            start_text.status = STARTED
            start_text.setAutoDraw(True)
        
        # if start_text is active this frame...
        if start_text.status == STARTED:
            # update params
            pass
        # *start_button* updates
        
        # if start_button is starting this frame...
        if start_button.status == NOT_STARTED and tThisFlip >= 0-frameTolerance:
            # keep track of start time/frame for later
            start_button.frameNStart = frameN  # exact frame index
            start_button.tStart = t  # local t and not account for scr refresh
            start_button.tStartRefresh = tThisFlipGlobal  # on global time
            win.timeOnFlip(start_button, 'tStartRefresh')  # time at next scr refresh
            # add timestamp to datafile
            thisExp.timestampOnFlip(win, 'start_button.started')
            # update status
            start_button.status = STARTED
            win.callOnFlip(start_button.buttonClock.reset)
            start_button.setAutoDraw(True)
        
        # if start_button is active this frame...
        if start_button.status == STARTED:
            # update params
            pass
            # check whether start_button has been pressed
            if start_button.isClicked:
                if not start_button.wasClicked:
                    # if this is a new click, store time of first click and clicked until
                    start_button.timesOn.append(start_button.buttonClock.getTime())
                    start_button.timesOff.append(start_button.buttonClock.getTime())
                elif len(start_button.timesOff):
                    # if click is continuing from last frame, update time of clicked until
                    start_button.timesOff[-1] = start_button.buttonClock.getTime()
                if not start_button.wasClicked:
                    # end routine when start_button is clicked
                    continueRoutine = False
                if not start_button.wasClicked:
                    # run callback code when start_button is clicked
                    pass
        # take note of whether start_button was clicked, so that next frame we know if clicks are new
        start_button.wasClicked = start_button.isClicked and start_button.status == STARTED
        
        # check for quit (typically the Esc key)
        if defaultKeyboard.getKeys(keyList=["escape"]):
            thisExp.status = FINISHED
        if thisExp.status == FINISHED or endExpNow:
            endExperiment(thisExp, win=win)
            return
        # pause experiment here if requested
        if thisExp.status == PAUSED:
            pauseExperiment(
                thisExp=thisExp, 
                win=win, 
                timers=[routineTimer, globalClock], 
                currentRoutine=rr_start,
            )
            # skip the frame we paused on
            continue
        
        # has a Component requested the Routine to end?
        if not continueRoutine:
            rr_start.forceEnded = routineForceEnded = True
        # has the Routine been forcibly ended?
        if rr_start.forceEnded or routineForceEnded:
            break
        # has every Component finished?
        continueRoutine = False
        for thisComponent in rr_start.components:
            if hasattr(thisComponent, "status") and thisComponent.status != FINISHED:
                continueRoutine = True
                break  # at least one component has not yet finished
        
        # refresh the screen
        if continueRoutine:  # don't flip if this routine is over or we'll get a blank screen
            win.flip()
    
    # --- Ending Routine "rr_start" ---
    for thisComponent in rr_start.components:
        if hasattr(thisComponent, "setAutoDraw"):
            thisComponent.setAutoDraw(False)
    # store stop times for rr_start
    rr_start.tStop = globalClock.getTime(format='float')
    rr_start.tStopRefresh = tThisFlipGlobal
    thisExp.addData('rr_start.stopped', rr_start.tStop)
    thisExp.addData('start_button.numClicks', start_button.numClicks)
    if start_button.numClicks:
       thisExp.addData('start_button.timesOn', start_button.timesOn)
       thisExp.addData('start_button.timesOff', start_button.timesOff)
    else:
       thisExp.addData('start_button.timesOn', "")
       thisExp.addData('start_button.timesOff', "")
    thisExp.nextEntry()
    # the Routine "rr_start" was not non-slip safe, so reset the non-slip timer
    routineTimer.reset()
    
    # set up handler to look after randomisation of conditions etc
    oo_mvc_type = data.TrialHandler2(
        name='oo_mvc_type',
        nReps=1.0, 
        method='sequential', 
        extraInfo=expInfo, 
        originPath=-1, 
        trialList=data.importConditions('oo_mvc_type.xlsx'), 
        seed=None, 
        isTrials=True, 
    )
    thisExp.addLoop(oo_mvc_type)  # add the loop to the experiment
    thisOo_mvc_type = oo_mvc_type.trialList[0]  # so we can initialise stimuli with some values
    # abbreviate parameter names if possible (e.g. rgb = thisOo_mvc_type.rgb)
    if thisOo_mvc_type != None:
        for paramName in thisOo_mvc_type:
            globals()[paramName] = thisOo_mvc_type[paramName]
    if thisSession is not None:
        # if running in a Session with a Liaison client, send data up to now
        thisSession.sendExperimentData()
    
    for thisOo_mvc_type in oo_mvc_type:
        oo_mvc_type.status = STARTED
        if hasattr(thisOo_mvc_type, 'status'):
            thisOo_mvc_type.status = STARTED
        currentLoop = oo_mvc_type
        thisExp.timestampOnFlip(win, 'thisRow.t', format=globalClock.format)
        if thisSession is not None:
            # if running in a Session with a Liaison client, send data up to now
            thisSession.sendExperimentData()
        # abbreviate parameter names if possible (e.g. rgb = thisOo_mvc_type.rgb)
        if thisOo_mvc_type != None:
            for paramName in thisOo_mvc_type:
                globals()[paramName] = thisOo_mvc_type[paramName]
        
        # set up handler to look after randomisation of conditions etc
        oo_mvc_rep = data.TrialHandler2(
            name='oo_mvc_rep',
            nReps=cond_num_maxforce, 
            method='sequential', 
            extraInfo=expInfo, 
            originPath=-1, 
            trialList=[None], 
            seed=None, 
            isTrials=True, 
        )
        thisExp.addLoop(oo_mvc_rep)  # add the loop to the experiment
        thisOo_mvc_rep = oo_mvc_rep.trialList[0]  # so we can initialise stimuli with some values
        # abbreviate parameter names if possible (e.g. rgb = thisOo_mvc_rep.rgb)
        if thisOo_mvc_rep != None:
            for paramName in thisOo_mvc_rep:
                globals()[paramName] = thisOo_mvc_rep[paramName]
        if thisSession is not None:
            # if running in a Session with a Liaison client, send data up to now
            thisSession.sendExperimentData()
        
        for thisOo_mvc_rep in oo_mvc_rep:
            oo_mvc_rep.status = STARTED
            if hasattr(thisOo_mvc_rep, 'status'):
                thisOo_mvc_rep.status = STARTED
            currentLoop = oo_mvc_rep
            thisExp.timestampOnFlip(win, 'thisRow.t', format=globalClock.format)
            if thisSession is not None:
                # if running in a Session with a Liaison client, send data up to now
                thisSession.sendExperimentData()
            # abbreviate parameter names if possible (e.g. rgb = thisOo_mvc_rep.rgb)
            if thisOo_mvc_rep != None:
                for paramName in thisOo_mvc_rep:
                    globals()[paramName] = thisOo_mvc_rep[paramName]
            
            # --- Prepare to start Routine "rr_countdown" ---
            # create an object to store info about Routine rr_countdown
            rr_countdown = data.Routine(
                name='rr_countdown',
                components=[countdown_lights, countdown_display],
            )
            rr_countdown.status = NOT_STARTED
            continueRoutine = True
            # update component parameters for each repeat
            # Run 'Begin Routine' code from countdown_code
            # Countdown setup
            cur_countdown_idx = len(INNER_HOLES)
            cur_countdown_last_update_sec = None
            
            # DEBUG
            countdown_lights.turnOffLights("all")
            # store start times for rr_countdown
            rr_countdown.tStartRefresh = win.getFutureFlipTime(clock=globalClock)
            rr_countdown.tStart = globalClock.getTime(format='float')
            rr_countdown.status = STARTED
            thisExp.addData('rr_countdown.started', rr_countdown.tStart)
            rr_countdown.maxDuration = None
            # keep track of which components have finished
            rr_countdownComponents = rr_countdown.components
            for thisComponent in rr_countdown.components:
                thisComponent.tStart = None
                thisComponent.tStop = None
                thisComponent.tStartRefresh = None
                thisComponent.tStopRefresh = None
                if hasattr(thisComponent, 'status'):
                    thisComponent.status = NOT_STARTED
            # reset timers
            t = 0
            _timeToFirstFrame = win.getFutureFlipTime(clock="now")
            frameN = -1
            
            # --- Run Routine "rr_countdown" ---
            thisExp.currentRoutine = rr_countdown
            rr_countdown.forceEnded = routineForceEnded = not continueRoutine
            while continueRoutine:
                # if trial has changed, end Routine now
                if hasattr(thisOo_mvc_rep, 'status') and thisOo_mvc_rep.status == STOPPING:
                    continueRoutine = False
                # get current time
                t = routineTimer.getTime()
                tThisFlip = win.getFutureFlipTime(clock=routineTimer)
                tThisFlipGlobal = win.getFutureFlipTime(clock=None)
                frameN = frameN + 1  # number of completed frames (so 0 is the first frame)
                # update/draw components on each frame
                # Run 'Each Frame' code from countdown_code
                # Check if it's time to update the countdown
                if cur_countdown_last_update_sec is None:
                    cur_countdown_last_update_sec = t  # t is the routine timer
                
                if t - cur_countdown_last_update_sec >= COUNTDOWN_STEP_SEC:
                    if cur_countdown_idx >= 0:
                        # Turn off the current light
                        countdown_lights.turnOffLights(cur_countdown_idx)
                    else:
                        continueRoutine = False
                    
                    # Decrease the number of active lights
                    cur_countdown_idx -= 1
                    cur_countdown_last_update_sec = t
                
                # if countdown_lights is starting this frame...
                if countdown_lights.status == NOT_STARTED and t >= 0-frameTolerance:
                    # keep track of start time/frame for later
                    countdown_lights.frameNStart = frameN  # exact frame index
                    countdown_lights.tStart = t  # local t and not account for scr refresh
                    countdown_lights.tStartRefresh = tThisFlipGlobal  # on global time
                    win.timeOnFlip(countdown_lights, 'tStartRefresh')  # time at next scr refresh
                    # add timestamp to datafile
                    thisExp.addData('countdown_lights.started', t)
                    # update status
                    countdown_lights.status = STARTED
                    countdown_lights.setLights("inner", 'white')
                
                # if countdown_lights is active this frame...
                if countdown_lights.status == STARTED:
                    # update params
                    pass
                    
                
                # if countdown_lights is stopping this frame...
                if countdown_lights.status == STARTED:
                    # is it time to stop? (based on global clock, using actual start)
                    if tThisFlipGlobal > countdown_lights.tStartRefresh + 4-frameTolerance:
                        # keep track of stop time/frame for later
                        countdown_lights.tStop = t  # not accounting for scr refresh
                        countdown_lights.tStopRefresh = tThisFlipGlobal  # on global time
                        countdown_lights.frameNStop = frameN  # exact frame index
                        # add timestamp to datafile
                        thisExp.addData('countdown_lights.stopped', t)
                        # update status
                        countdown_lights.status = FINISHED
                        if False:
                           countdown_lights.turnOffLights("inner")
                
                # *countdown_display* updates
                
                # if countdown_display is starting this frame...
                if countdown_display.status == NOT_STARTED and tThisFlip >= 0.0-frameTolerance:
                    # keep track of start time/frame for later
                    countdown_display.frameNStart = frameN  # exact frame index
                    countdown_display.tStart = t  # local t and not account for scr refresh
                    countdown_display.tStartRefresh = tThisFlipGlobal  # on global time
                    win.timeOnFlip(countdown_display, 'tStartRefresh')  # time at next scr refresh
                    # add timestamp to datafile
                    thisExp.timestampOnFlip(win, 'countdown_display.started')
                    # update status
                    countdown_display.status = STARTED
                    countdown_display.setAutoDraw(True)
                
                # if countdown_display is active this frame...
                if countdown_display.status == STARTED:
                    # update params
                    countdown_display.setText(cur_countdown_idx + 1, log=False)
                
                # check for quit (typically the Esc key)
                if defaultKeyboard.getKeys(keyList=["escape"]):
                    thisExp.status = FINISHED
                if thisExp.status == FINISHED or endExpNow:
                    endExperiment(thisExp, win=win)
                    return
                # pause experiment here if requested
                if thisExp.status == PAUSED:
                    pauseExperiment(
                        thisExp=thisExp, 
                        win=win, 
                        timers=[routineTimer, globalClock], 
                        currentRoutine=rr_countdown,
                    )
                    # skip the frame we paused on
                    continue
                
                # has a Component requested the Routine to end?
                if not continueRoutine:
                    rr_countdown.forceEnded = routineForceEnded = True
                # has the Routine been forcibly ended?
                if rr_countdown.forceEnded or routineForceEnded:
                    break
                # has every Component finished?
                continueRoutine = False
                for thisComponent in rr_countdown.components:
                    if hasattr(thisComponent, "status") and thisComponent.status != FINISHED:
                        continueRoutine = True
                        break  # at least one component has not yet finished
                
                # refresh the screen
                if continueRoutine:  # don't flip if this routine is over or we'll get a blank screen
                    win.flip()
            
            # --- Ending Routine "rr_countdown" ---
            for thisComponent in rr_countdown.components:
                if hasattr(thisComponent, "setAutoDraw"):
                    thisComponent.setAutoDraw(False)
            # store stop times for rr_countdown
            rr_countdown.tStop = globalClock.getTime(format='float')
            rr_countdown.tStopRefresh = tThisFlipGlobal
            thisExp.addData('rr_countdown.stopped', rr_countdown.tStop)
            if False:
                countdown_lights.turnOffLights("inner")
            # the Routine "rr_countdown" was not non-slip safe, so reset the non-slip timer
            routineTimer.reset()
            
            # --- Prepare to start Routine "rr_maxforce" ---
            # create an object to store info about Routine rr_maxforce
            rr_maxforce = data.Routine(
                name='rr_maxforce',
                components=[maxforce_lights, maxforce_force, maxforce_display],
            )
            rr_maxforce.status = NOT_STARTED
            continueRoutine = True
            # update component parameters for each repeat
            # Run 'Begin Routine' code from maxforce_code
            # Initialize current value of MVC
            cur_mvc = maxforce_force.whiteForce = 0
            
            # Improve display appearance
            maxforce_display.alignText = "left"
            maxforce_display.bold = True
            # store start times for rr_maxforce
            rr_maxforce.tStartRefresh = win.getFutureFlipTime(clock=globalClock)
            rr_maxforce.tStart = globalClock.getTime(format='float')
            rr_maxforce.status = STARTED
            thisExp.addData('rr_maxforce.started', rr_maxforce.tStart)
            rr_maxforce.maxDuration = None
            # keep track of which components have finished
            rr_maxforceComponents = rr_maxforce.components
            for thisComponent in rr_maxforce.components:
                thisComponent.tStart = None
                thisComponent.tStop = None
                thisComponent.tStartRefresh = None
                thisComponent.tStopRefresh = None
                if hasattr(thisComponent, 'status'):
                    thisComponent.status = NOT_STARTED
            # reset timers
            t = 0
            _timeToFirstFrame = win.getFutureFlipTime(clock="now")
            frameN = -1
            
            # --- Run Routine "rr_maxforce" ---
            thisExp.currentRoutine = rr_maxforce
            rr_maxforce.forceEnded = routineForceEnded = not continueRoutine
            while continueRoutine and routineTimer.getTime() < 3.5:
                # if trial has changed, end Routine now
                if hasattr(thisOo_mvc_rep, 'status') and thisOo_mvc_rep.status == STOPPING:
                    continueRoutine = False
                # get current time
                t = routineTimer.getTime()
                tThisFlip = win.getFutureFlipTime(clock=routineTimer)
                tThisFlipGlobal = win.getFutureFlipTime(clock=None)
                frameN = frameN + 1  # number of completed frames (so 0 is the first frame)
                # update/draw components on each frame
                # Run 'Each Frame' code from maxforce_code
                # Update current MVC value based on dynamometer measurement
                cur_mvc = max(cur_mvc, maxforce_force.whiteForce)
                
                # Update MVC only if we are in a calibration trial
                if cond_maxforce_type == "calibration":
                    mvc = max(mvc, cur_mvc)
                
                # Display text
                maxforce_display_text = (
                    f"============ TRIAL INFO ============\n"
                    f"Trial Type:        {cond_maxforce_type}\n"
                    f"Current Force:     {maxforce_force.whiteForce:.2f} N\n"
                    f"MVC (cur / max):   {cur_mvc:.2f} / {mvc:.2f} N\n"
                    f"====================================\n"
                )
                
                # if maxforce_lights is starting this frame...
                if maxforce_lights.status == NOT_STARTED and t >= 0.5-frameTolerance:
                    # keep track of start time/frame for later
                    maxforce_lights.frameNStart = frameN  # exact frame index
                    maxforce_lights.tStart = t  # local t and not account for scr refresh
                    maxforce_lights.tStartRefresh = tThisFlipGlobal  # on global time
                    win.timeOnFlip(maxforce_lights, 'tStartRefresh')  # time at next scr refresh
                    # add timestamp to datafile
                    thisExp.addData('maxforce_lights.started', t)
                    # update status
                    maxforce_lights.status = STARTED
                    maxforce_lights.setLights("all", 'white')
                
                # if maxforce_lights is active this frame...
                if maxforce_lights.status == STARTED:
                    # update params
                    pass
                    
                
                # if maxforce_lights is stopping this frame...
                if maxforce_lights.status == STARTED:
                    # is it time to stop? (based on global clock, using actual start)
                    if tThisFlipGlobal > maxforce_lights.tStartRefresh + 3.0-frameTolerance:
                        # keep track of stop time/frame for later
                        maxforce_lights.tStop = t  # not accounting for scr refresh
                        maxforce_lights.tStopRefresh = tThisFlipGlobal  # on global time
                        maxforce_lights.frameNStop = frameN  # exact frame index
                        # add timestamp to datafile
                        thisExp.addData('maxforce_lights.stopped', t)
                        # update status
                        maxforce_lights.status = FINISHED
                        if True:
                           maxforce_lights.turnOffLights("all")
                
                # if maxforce_force is starting this frame...
                if maxforce_force.status == NOT_STARTED and t >= 0.5-frameTolerance:
                    # keep track of start time/frame for later
                    maxforce_force.frameNStart = frameN  # exact frame index
                    maxforce_force.tStart = t  # local t and not account for scr refresh
                    maxforce_force.tStartRefresh = tThisFlipGlobal  # on global time
                    win.timeOnFlip(maxforce_force, 'tStartRefresh')  # time at next scr refresh
                    # add timestamp to datafile
                    thisExp.addData('maxforce_force.started', t)
                    # update status
                    maxforce_force.status = STARTED
                    maxforce_force.startForceMeasurement(60, 'both')
                
                # if maxforce_force is active this frame...
                if maxforce_force.status == STARTED:
                    # update params
                    pass
                    maxforce_force.updateForceMeasurement()
                
                # if maxforce_force is stopping this frame...
                if maxforce_force.status == STARTED:
                    # is it time to stop? (based on global clock, using actual start)
                    if tThisFlipGlobal > maxforce_force.tStartRefresh + 3-frameTolerance:
                        # keep track of stop time/frame for later
                        maxforce_force.tStop = t  # not accounting for scr refresh
                        maxforce_force.tStopRefresh = tThisFlipGlobal  # on global time
                        maxforce_force.frameNStop = frameN  # exact frame index
                        # add timestamp to datafile
                        thisExp.addData('maxforce_force.stopped', t)
                        # update status
                        maxforce_force.status = FINISHED
                        maxforce_force.stopForceMeasurement()
                
                # *maxforce_display* updates
                
                # if maxforce_display is starting this frame...
                if maxforce_display.status == NOT_STARTED and tThisFlip >= 0-frameTolerance:
                    # keep track of start time/frame for later
                    maxforce_display.frameNStart = frameN  # exact frame index
                    maxforce_display.tStart = t  # local t and not account for scr refresh
                    maxforce_display.tStartRefresh = tThisFlipGlobal  # on global time
                    win.timeOnFlip(maxforce_display, 'tStartRefresh')  # time at next scr refresh
                    # add timestamp to datafile
                    thisExp.timestampOnFlip(win, 'maxforce_display.started')
                    # update status
                    maxforce_display.status = STARTED
                    maxforce_display.setAutoDraw(True)
                
                # if maxforce_display is active this frame...
                if maxforce_display.status == STARTED:
                    # update params
                    maxforce_display.setText(maxforce_display_text, log=False)
                
                # if maxforce_display is stopping this frame...
                if maxforce_display.status == STARTED:
                    # is it time to stop? (based on global clock, using actual start)
                    if tThisFlipGlobal > maxforce_display.tStartRefresh + 3-frameTolerance:
                        # keep track of stop time/frame for later
                        maxforce_display.tStop = t  # not accounting for scr refresh
                        maxforce_display.tStopRefresh = tThisFlipGlobal  # on global time
                        maxforce_display.frameNStop = frameN  # exact frame index
                        # add timestamp to datafile
                        thisExp.timestampOnFlip(win, 'maxforce_display.stopped')
                        # update status
                        maxforce_display.status = FINISHED
                        maxforce_display.setAutoDraw(False)
                
                # check for quit (typically the Esc key)
                if defaultKeyboard.getKeys(keyList=["escape"]):
                    thisExp.status = FINISHED
                if thisExp.status == FINISHED or endExpNow:
                    endExperiment(thisExp, win=win)
                    return
                # pause experiment here if requested
                if thisExp.status == PAUSED:
                    pauseExperiment(
                        thisExp=thisExp, 
                        win=win, 
                        timers=[routineTimer, globalClock], 
                        currentRoutine=rr_maxforce,
                    )
                    # skip the frame we paused on
                    continue
                
                # has a Component requested the Routine to end?
                if not continueRoutine:
                    rr_maxforce.forceEnded = routineForceEnded = True
                # has the Routine been forcibly ended?
                if rr_maxforce.forceEnded or routineForceEnded:
                    break
                # has every Component finished?
                continueRoutine = False
                for thisComponent in rr_maxforce.components:
                    if hasattr(thisComponent, "status") and thisComponent.status != FINISHED:
                        continueRoutine = True
                        break  # at least one component has not yet finished
                
                # refresh the screen
                if continueRoutine:  # don't flip if this routine is over or we'll get a blank screen
                    win.flip()
            
            # --- Ending Routine "rr_maxforce" ---
            for thisComponent in rr_maxforce.components:
                if hasattr(thisComponent, "setAutoDraw"):
                    thisComponent.setAutoDraw(False)
            # store stop times for rr_maxforce
            rr_maxforce.tStop = globalClock.getTime(format='float')
            rr_maxforce.tStopRefresh = tThisFlipGlobal
            thisExp.addData('rr_maxforce.stopped', rr_maxforce.tStop)
            if True:
                maxforce_lights.turnOffLights("all")
            oo_mvc_rep.addData('maxforce_force.rate', 60)
            oo_mvc_rep.addData('maxforce_force.dynamometer', 'both')
            oo_mvc_rep.addData('maxforce_force.maxWhiteForce', maxforce_force.maxWhiteForce)
            oo_mvc_rep.addData('maxforce_force.maxBlueForce', maxforce_force.maxBlueForce)
            if True:
                _raw_path = thisExp.dataFileName + '_force_long.tsv'
                _write_header = not os.path.exists(_raw_path)
                _loop = oo_mvc_rep
                _trial_index = _loop.thisN if _loop is not None and hasattr(_loop, 'thisN') else -1
                _trial_name = _loop.name if _loop is not None and hasattr(_loop, 'name') else ''
                _identifier = str("test") if "test" is not None else ''
                _records = maxforce_force.forceRows if hasattr(maxforce_force, 'forceRows') else []
                _times = maxforce_force.times
                _white = maxforce_force.whiteForceValues
                _blue = maxforce_force.blueForceValues
                _n = max(len(_times), len(_white), len(_blue))
                with open(_raw_path, 'a', encoding='utf-8') as _f:
                    if _write_header:
                        _f.write('participant	session	routine	component	trial_index	trial_name	identifier	sample_index	white_time	blue_time	time	white_force	blue_force	white_force_raw_counts	blue_force_raw_counts\n')
                    for _i, _record in enumerate(_records):
                        _row = [
                            expInfo.get("participant", ""),
                            expInfo.get("session", ""),
                            'rr_maxforce',
                            'maxforce_force',
                            _trial_index,
                            _trial_name,
                            _identifier,
                            _i,
                            _record['white_time'],
                            _record['blue_time'],
                            _record['time'],
                            _record['white_force'],
                            _record['blue_force'],
                            _record['white_force_raw_counts'] if _record['white_force_raw_counts'] is not None else '',
                            _record['blue_force_raw_counts'] if _record['blue_force_raw_counts'] is not None else '',
                        ]
                        _f.write('	'.join(str(_v) for _v in _row) + '\n')
            # using non-slip timing so subtract the expected duration of this Routine (unless ended on request)
            if rr_maxforce.maxDurationReached:
                routineTimer.addTime(-rr_maxforce.maxDuration)
            elif rr_maxforce.forceEnded:
                routineTimer.reset()
            else:
                routineTimer.addTime(-3.500000)
            
            # --- Prepare to start Routine "rr_mvc" ---
            # create an object to store info about Routine rr_mvc
            rr_mvc = data.Routine(
                name='rr_mvc',
                components=[mvc_display, mvc_continue],
            )
            rr_mvc.status = NOT_STARTED
            continueRoutine = True
            # update component parameters for each repeat
            # Run 'Begin Routine' code from mvc_code
            # Improve display appearance
            mvc_display.alignText = "left"
            mvc_display.bold = True
            # reset mvc_continue to account for continued clicks & clear times on/off
            mvc_continue.reset()
            # store start times for rr_mvc
            rr_mvc.tStartRefresh = win.getFutureFlipTime(clock=globalClock)
            rr_mvc.tStart = globalClock.getTime(format='float')
            rr_mvc.status = STARTED
            thisExp.addData('rr_mvc.started', rr_mvc.tStart)
            rr_mvc.maxDuration = None
            # keep track of which components have finished
            rr_mvcComponents = rr_mvc.components
            for thisComponent in rr_mvc.components:
                thisComponent.tStart = None
                thisComponent.tStop = None
                thisComponent.tStartRefresh = None
                thisComponent.tStopRefresh = None
                if hasattr(thisComponent, 'status'):
                    thisComponent.status = NOT_STARTED
            # reset timers
            t = 0
            _timeToFirstFrame = win.getFutureFlipTime(clock="now")
            frameN = -1
            
            # --- Run Routine "rr_mvc" ---
            thisExp.currentRoutine = rr_mvc
            rr_mvc.forceEnded = routineForceEnded = not continueRoutine
            while continueRoutine:
                # if trial has changed, end Routine now
                if hasattr(thisOo_mvc_rep, 'status') and thisOo_mvc_rep.status == STOPPING:
                    continueRoutine = False
                # get current time
                t = routineTimer.getTime()
                tThisFlip = win.getFutureFlipTime(clock=routineTimer)
                tThisFlipGlobal = win.getFutureFlipTime(clock=None)
                frameN = frameN + 1  # number of completed frames (so 0 is the first frame)
                # update/draw components on each frame
                
                # *mvc_display* updates
                
                # if mvc_display is starting this frame...
                if mvc_display.status == NOT_STARTED and tThisFlip >= 0.0-frameTolerance:
                    # keep track of start time/frame for later
                    mvc_display.frameNStart = frameN  # exact frame index
                    mvc_display.tStart = t  # local t and not account for scr refresh
                    mvc_display.tStartRefresh = tThisFlipGlobal  # on global time
                    win.timeOnFlip(mvc_display, 'tStartRefresh')  # time at next scr refresh
                    # add timestamp to datafile
                    thisExp.timestampOnFlip(win, 'mvc_display.started')
                    # update status
                    mvc_display.status = STARTED
                    mvc_display.setAutoDraw(True)
                
                # if mvc_display is active this frame...
                if mvc_display.status == STARTED:
                    # update params
                    mvc_display.setText(maxforce_display_text, log=False)
                # *mvc_continue* updates
                
                # if mvc_continue is starting this frame...
                if mvc_continue.status == NOT_STARTED and tThisFlip >= 0-frameTolerance:
                    # keep track of start time/frame for later
                    mvc_continue.frameNStart = frameN  # exact frame index
                    mvc_continue.tStart = t  # local t and not account for scr refresh
                    mvc_continue.tStartRefresh = tThisFlipGlobal  # on global time
                    win.timeOnFlip(mvc_continue, 'tStartRefresh')  # time at next scr refresh
                    # add timestamp to datafile
                    thisExp.timestampOnFlip(win, 'mvc_continue.started')
                    # update status
                    mvc_continue.status = STARTED
                    win.callOnFlip(mvc_continue.buttonClock.reset)
                    mvc_continue.setAutoDraw(True)
                
                # if mvc_continue is active this frame...
                if mvc_continue.status == STARTED:
                    # update params
                    pass
                    # check whether mvc_continue has been pressed
                    if mvc_continue.isClicked:
                        if not mvc_continue.wasClicked:
                            # if this is a new click, store time of first click and clicked until
                            mvc_continue.timesOn.append(mvc_continue.buttonClock.getTime())
                            mvc_continue.timesOff.append(mvc_continue.buttonClock.getTime())
                        elif len(mvc_continue.timesOff):
                            # if click is continuing from last frame, update time of clicked until
                            mvc_continue.timesOff[-1] = mvc_continue.buttonClock.getTime()
                        if not mvc_continue.wasClicked:
                            # end routine when mvc_continue is clicked
                            continueRoutine = False
                        if not mvc_continue.wasClicked:
                            # run callback code when mvc_continue is clicked
                            pass
                # take note of whether mvc_continue was clicked, so that next frame we know if clicks are new
                mvc_continue.wasClicked = mvc_continue.isClicked and mvc_continue.status == STARTED
                
                # check for quit (typically the Esc key)
                if defaultKeyboard.getKeys(keyList=["escape"]):
                    thisExp.status = FINISHED
                if thisExp.status == FINISHED or endExpNow:
                    endExperiment(thisExp, win=win)
                    return
                # pause experiment here if requested
                if thisExp.status == PAUSED:
                    pauseExperiment(
                        thisExp=thisExp, 
                        win=win, 
                        timers=[routineTimer, globalClock], 
                        currentRoutine=rr_mvc,
                    )
                    # skip the frame we paused on
                    continue
                
                # has a Component requested the Routine to end?
                if not continueRoutine:
                    rr_mvc.forceEnded = routineForceEnded = True
                # has the Routine been forcibly ended?
                if rr_mvc.forceEnded or routineForceEnded:
                    break
                # has every Component finished?
                continueRoutine = False
                for thisComponent in rr_mvc.components:
                    if hasattr(thisComponent, "status") and thisComponent.status != FINISHED:
                        continueRoutine = True
                        break  # at least one component has not yet finished
                
                # refresh the screen
                if continueRoutine:  # don't flip if this routine is over or we'll get a blank screen
                    win.flip()
            
            # --- Ending Routine "rr_mvc" ---
            for thisComponent in rr_mvc.components:
                if hasattr(thisComponent, "setAutoDraw"):
                    thisComponent.setAutoDraw(False)
            # store stop times for rr_mvc
            rr_mvc.tStop = globalClock.getTime(format='float')
            rr_mvc.tStopRefresh = tThisFlipGlobal
            thisExp.addData('rr_mvc.stopped', rr_mvc.tStop)
            oo_mvc_rep.addData('mvc_continue.numClicks', mvc_continue.numClicks)
            if mvc_continue.numClicks:
               oo_mvc_rep.addData('mvc_continue.timesOn', mvc_continue.timesOn)
               oo_mvc_rep.addData('mvc_continue.timesOff', mvc_continue.timesOff)
            else:
               oo_mvc_rep.addData('mvc_continue.timesOn', "")
               oo_mvc_rep.addData('mvc_continue.timesOff', "")
            # the Routine "rr_mvc" was not non-slip safe, so reset the non-slip timer
            routineTimer.reset()
            # mark thisOo_mvc_rep as finished
            if hasattr(thisOo_mvc_rep, 'status'):
                thisOo_mvc_rep.status = FINISHED
            # if awaiting a pause, pause now
            if oo_mvc_rep.status == PAUSED:
                thisExp.status = PAUSED
                pauseExperiment(
                    thisExp=thisExp, 
                    win=win, 
                    timers=[globalClock], 
                )
                # once done pausing, restore running status
                oo_mvc_rep.status = STARTED
            thisExp.nextEntry()
            
        # completed cond_num_maxforce repeats of 'oo_mvc_rep'
        oo_mvc_rep.status = FINISHED
        
        if thisSession is not None:
            # if running in a Session with a Liaison client, send data up to now
            thisSession.sendExperimentData()
        # mark thisOo_mvc_type as finished
        if hasattr(thisOo_mvc_type, 'status'):
            thisOo_mvc_type.status = FINISHED
        # if awaiting a pause, pause now
        if oo_mvc_type.status == PAUSED:
            thisExp.status = PAUSED
            pauseExperiment(
                thisExp=thisExp, 
                win=win, 
                timers=[globalClock], 
            )
            # once done pausing, restore running status
            oo_mvc_type.status = STARTED
        thisExp.nextEntry()
        
    # completed 1.0 repeats of 'oo_mvc_type'
    oo_mvc_type.status = FINISHED
    
    if thisSession is not None:
        # if running in a Session with a Liaison client, send data up to now
        thisSession.sendExperimentData()
    
    # set up handler to look after randomisation of conditions etc
    oo_block = data.TrialHandler2(
        name='oo_block',
        nReps=NUM_BLOCKS, 
        method='random', 
        extraInfo=expInfo, 
        originPath=-1, 
        trialList=[None], 
        seed=None, 
        isTrials=True, 
    )
    thisExp.addLoop(oo_block)  # add the loop to the experiment
    thisOo_block = oo_block.trialList[0]  # so we can initialise stimuli with some values
    # abbreviate parameter names if possible (e.g. rgb = thisOo_block.rgb)
    if thisOo_block != None:
        for paramName in thisOo_block:
            globals()[paramName] = thisOo_block[paramName]
    if thisSession is not None:
        # if running in a Session with a Liaison client, send data up to now
        thisSession.sendExperimentData()
    
    for thisOo_block in oo_block:
        oo_block.status = STARTED
        if hasattr(thisOo_block, 'status'):
            thisOo_block.status = STARTED
        currentLoop = oo_block
        thisExp.timestampOnFlip(win, 'thisRow.t', format=globalClock.format)
        if thisSession is not None:
            # if running in a Session with a Liaison client, send data up to now
            thisSession.sendExperimentData()
        # abbreviate parameter names if possible (e.g. rgb = thisOo_block.rgb)
        if thisOo_block != None:
            for paramName in thisOo_block:
                globals()[paramName] = thisOo_block[paramName]
        
        # --- Prepare to start Routine "rr_countdown" ---
        # create an object to store info about Routine rr_countdown
        rr_countdown = data.Routine(
            name='rr_countdown',
            components=[countdown_lights, countdown_display],
        )
        rr_countdown.status = NOT_STARTED
        continueRoutine = True
        # update component parameters for each repeat
        # Run 'Begin Routine' code from countdown_code
        # Countdown setup
        cur_countdown_idx = len(INNER_HOLES)
        cur_countdown_last_update_sec = None
        
        # DEBUG
        countdown_lights.turnOffLights("all")
        # store start times for rr_countdown
        rr_countdown.tStartRefresh = win.getFutureFlipTime(clock=globalClock)
        rr_countdown.tStart = globalClock.getTime(format='float')
        rr_countdown.status = STARTED
        thisExp.addData('rr_countdown.started', rr_countdown.tStart)
        rr_countdown.maxDuration = None
        # keep track of which components have finished
        rr_countdownComponents = rr_countdown.components
        for thisComponent in rr_countdown.components:
            thisComponent.tStart = None
            thisComponent.tStop = None
            thisComponent.tStartRefresh = None
            thisComponent.tStopRefresh = None
            if hasattr(thisComponent, 'status'):
                thisComponent.status = NOT_STARTED
        # reset timers
        t = 0
        _timeToFirstFrame = win.getFutureFlipTime(clock="now")
        frameN = -1
        
        # --- Run Routine "rr_countdown" ---
        thisExp.currentRoutine = rr_countdown
        rr_countdown.forceEnded = routineForceEnded = not continueRoutine
        while continueRoutine:
            # if trial has changed, end Routine now
            if hasattr(thisOo_block, 'status') and thisOo_block.status == STOPPING:
                continueRoutine = False
            # get current time
            t = routineTimer.getTime()
            tThisFlip = win.getFutureFlipTime(clock=routineTimer)
            tThisFlipGlobal = win.getFutureFlipTime(clock=None)
            frameN = frameN + 1  # number of completed frames (so 0 is the first frame)
            # update/draw components on each frame
            # Run 'Each Frame' code from countdown_code
            # Check if it's time to update the countdown
            if cur_countdown_last_update_sec is None:
                cur_countdown_last_update_sec = t  # t is the routine timer
            
            if t - cur_countdown_last_update_sec >= COUNTDOWN_STEP_SEC:
                if cur_countdown_idx >= 0:
                    # Turn off the current light
                    countdown_lights.turnOffLights(cur_countdown_idx)
                else:
                    continueRoutine = False
                
                # Decrease the number of active lights
                cur_countdown_idx -= 1
                cur_countdown_last_update_sec = t
            
            # if countdown_lights is starting this frame...
            if countdown_lights.status == NOT_STARTED and t >= 0-frameTolerance:
                # keep track of start time/frame for later
                countdown_lights.frameNStart = frameN  # exact frame index
                countdown_lights.tStart = t  # local t and not account for scr refresh
                countdown_lights.tStartRefresh = tThisFlipGlobal  # on global time
                win.timeOnFlip(countdown_lights, 'tStartRefresh')  # time at next scr refresh
                # add timestamp to datafile
                thisExp.addData('countdown_lights.started', t)
                # update status
                countdown_lights.status = STARTED
                countdown_lights.setLights("inner", 'white')
            
            # if countdown_lights is active this frame...
            if countdown_lights.status == STARTED:
                # update params
                pass
                
            
            # if countdown_lights is stopping this frame...
            if countdown_lights.status == STARTED:
                # is it time to stop? (based on global clock, using actual start)
                if tThisFlipGlobal > countdown_lights.tStartRefresh + 4-frameTolerance:
                    # keep track of stop time/frame for later
                    countdown_lights.tStop = t  # not accounting for scr refresh
                    countdown_lights.tStopRefresh = tThisFlipGlobal  # on global time
                    countdown_lights.frameNStop = frameN  # exact frame index
                    # add timestamp to datafile
                    thisExp.addData('countdown_lights.stopped', t)
                    # update status
                    countdown_lights.status = FINISHED
                    if False:
                       countdown_lights.turnOffLights("inner")
            
            # *countdown_display* updates
            
            # if countdown_display is starting this frame...
            if countdown_display.status == NOT_STARTED and tThisFlip >= 0.0-frameTolerance:
                # keep track of start time/frame for later
                countdown_display.frameNStart = frameN  # exact frame index
                countdown_display.tStart = t  # local t and not account for scr refresh
                countdown_display.tStartRefresh = tThisFlipGlobal  # on global time
                win.timeOnFlip(countdown_display, 'tStartRefresh')  # time at next scr refresh
                # add timestamp to datafile
                thisExp.timestampOnFlip(win, 'countdown_display.started')
                # update status
                countdown_display.status = STARTED
                countdown_display.setAutoDraw(True)
            
            # if countdown_display is active this frame...
            if countdown_display.status == STARTED:
                # update params
                countdown_display.setText(cur_countdown_idx + 1, log=False)
            
            # check for quit (typically the Esc key)
            if defaultKeyboard.getKeys(keyList=["escape"]):
                thisExp.status = FINISHED
            if thisExp.status == FINISHED or endExpNow:
                endExperiment(thisExp, win=win)
                return
            # pause experiment here if requested
            if thisExp.status == PAUSED:
                pauseExperiment(
                    thisExp=thisExp, 
                    win=win, 
                    timers=[routineTimer, globalClock], 
                    currentRoutine=rr_countdown,
                )
                # skip the frame we paused on
                continue
            
            # has a Component requested the Routine to end?
            if not continueRoutine:
                rr_countdown.forceEnded = routineForceEnded = True
            # has the Routine been forcibly ended?
            if rr_countdown.forceEnded or routineForceEnded:
                break
            # has every Component finished?
            continueRoutine = False
            for thisComponent in rr_countdown.components:
                if hasattr(thisComponent, "status") and thisComponent.status != FINISHED:
                    continueRoutine = True
                    break  # at least one component has not yet finished
            
            # refresh the screen
            if continueRoutine:  # don't flip if this routine is over or we'll get a blank screen
                win.flip()
        
        # --- Ending Routine "rr_countdown" ---
        for thisComponent in rr_countdown.components:
            if hasattr(thisComponent, "setAutoDraw"):
                thisComponent.setAutoDraw(False)
        # store stop times for rr_countdown
        rr_countdown.tStop = globalClock.getTime(format='float')
        rr_countdown.tStopRefresh = tThisFlipGlobal
        thisExp.addData('rr_countdown.stopped', rr_countdown.tStop)
        if False:
            countdown_lights.turnOffLights("inner")
        # the Routine "rr_countdown" was not non-slip safe, so reset the non-slip timer
        routineTimer.reset()
        
        # set up handler to look after randomisation of conditions etc
        oo_level = data.TrialHandler2(
            name='oo_level',
            nReps=NUM_LEVELS, 
            method='random', 
            extraInfo=expInfo, 
            originPath=-1, 
            trialList=data.importConditions('oo_level.xlsx'), 
            seed=None, 
            isTrials=True, 
        )
        thisExp.addLoop(oo_level)  # add the loop to the experiment
        thisOo_level = oo_level.trialList[0]  # so we can initialise stimuli with some values
        # abbreviate parameter names if possible (e.g. rgb = thisOo_level.rgb)
        if thisOo_level != None:
            for paramName in thisOo_level:
                globals()[paramName] = thisOo_level[paramName]
        if thisSession is not None:
            # if running in a Session with a Liaison client, send data up to now
            thisSession.sendExperimentData()
        
        for thisOo_level in oo_level:
            oo_level.status = STARTED
            if hasattr(thisOo_level, 'status'):
                thisOo_level.status = STARTED
            currentLoop = oo_level
            thisExp.timestampOnFlip(win, 'thisRow.t', format=globalClock.format)
            if thisSession is not None:
                # if running in a Session with a Liaison client, send data up to now
                thisSession.sendExperimentData()
            # abbreviate parameter names if possible (e.g. rgb = thisOo_level.rgb)
            if thisOo_level != None:
                for paramName in thisOo_level:
                    globals()[paramName] = thisOo_level[paramName]
            
            # --- Prepare to start Routine "rr_target" ---
            # create an object to store info about Routine rr_target
            rr_target = data.Routine(
                name='rr_target',
                components=[target_lights, target_force, target_display],
            )
            rr_target.status = NOT_STARTED
            continueRoutine = True
            # update component parameters for each repeat
            # Run 'Begin Routine' code from target_code
            # ===== CONDITION SETUP =====
            
            # Get data for current block and trial
            cur_physical_demand_level  = blocks[oo_block.thisN][oo_level.thisN]['physical_demand_level']
            cur_cognitive_demand_level = blocks[oo_block.thisN][oo_level.thisN]['cognitive_demand_level']
            cur_target_color_rgb255    = blocks[oo_block.thisN][oo_level.thisN]['target_color_rgb255']
            
            # ===== PERFORMANCE SETUP =====
            
            # Reset hits/misses/fa/cr for dprime calculation later in slider Routine
            hits = misses = false_alarms = correct_rejections = 0
            
            # =============================================================================
            # AUDIO - BEGIN ROUTINE
            # =============================================================================
            with _state_lock:
                _target_amp  = 0.0
                _target_freq = FREQ_LOW
            
            targetForce     = cur_physical_demand_level * mvc
            lower_threshold = max(targetForce - TOLERANCE, 0)
            upper_threshold = targetForce + TOLERANCE
            
            toneState    = 0
            _lastSwitchT = -1e9
            _last_beep_t = -1e9  # tracks last beep time
            
            
            # ===== DISPLAY SETUP =====
            
            target_display.alignText = "left"
            target_display.bold = True
            cur_target_color_name = COLOR_NAMES.get(tuple(cur_target_color_rgb255), "Unknown")
            
            # Color the display circle
            for i, circle in enumerate(hole_stimuli):
                circle.fillColor = cur_target_color_rgb255
            
            # ===== DATA DUMPING =====
            pd.DataFrame(rows).to_csv(thisExp.dataFileName + '_design.csv', index=False)
            # store start times for rr_target
            rr_target.tStartRefresh = win.getFutureFlipTime(clock=globalClock)
            rr_target.tStart = globalClock.getTime(format='float')
            rr_target.status = STARTED
            thisExp.addData('rr_target.started', rr_target.tStart)
            rr_target.maxDuration = None
            # keep track of which components have finished
            rr_targetComponents = rr_target.components
            for thisComponent in rr_target.components:
                thisComponent.tStart = None
                thisComponent.tStop = None
                thisComponent.tStartRefresh = None
                thisComponent.tStopRefresh = None
                if hasattr(thisComponent, 'status'):
                    thisComponent.status = NOT_STARTED
            # reset timers
            t = 0
            _timeToFirstFrame = win.getFutureFlipTime(clock="now")
            frameN = -1
            
            # --- Run Routine "rr_target" ---
            thisExp.currentRoutine = rr_target
            rr_target.forceEnded = routineForceEnded = not continueRoutine
            while continueRoutine and routineTimer.getTime() < 3.0:
                # if trial has changed, end Routine now
                if hasattr(thisOo_level, 'status') and thisOo_level.status == STOPPING:
                    continueRoutine = False
                # get current time
                t = routineTimer.getTime()
                tThisFlip = win.getFutureFlipTime(clock=routineTimer)
                tThisFlipGlobal = win.getFutureFlipTime(clock=None)
                frameN = frameN + 1  # number of completed frames (so 0 is the first frame)
                # update/draw components on each frame
                # Run 'Each Frame' code from target_code
                # =============================================================================
                # AUDIO - EACH FRAME
                # =============================================================================
                if target_force.whiteForce is not None:
                    if target_force.whiteForce > upper_threshold:
                        proposed = 2
                    elif target_force.whiteForce < lower_threshold:
                        proposed = 1
                    else:
                        proposed = 0
                
                    if proposed != toneState and (t - _lastSwitchT) >= STATE_HOLD:
                        toneState    = proposed
                        _lastSwitchT = t
                        if toneState == 2:
                            freq, amp = FREQ_HIGH, AMP_MAX
                        elif toneState == 1:
                            freq, amp = FREQ_LOW, AMP_MAX
                        else:
                            freq, amp = FREQ_LOW, 0.0
                        with _state_lock:
                            _target_freq = freq
                            _target_amp  = amp
                else:
                    print("Force is None!")
                
                # Create the target display text
                if target_force.whiteForce > upper_threshold:
                    deviation_symbol = "\u2191\u2191\u2191"  # ↑
                elif target_force.whiteForce < lower_threshold:
                    deviation_symbol = "\u2193\u2193\u2193"  # ↓
                else:
                    deviation_symbol = "\u2013\u2013\u2013"  # – (on target)
                
                target_display_text = (
                    f"============ TRIAL INFO ============\n"
                    f"Target Color:      {cur_target_color_name}\n"
                    f"Physical demand:   {cur_physical_demand_level}\n"
                    f"Cognitive demand:  {cur_cognitive_demand_level}\n"
                    f"------------------------------------\n"
                    f"MVC:               {mvc:>4.0f} N\n"
                    f"Target:            {lower_threshold:>3.0f} \u007C {targetForce:>3.0f} \u007C {upper_threshold:>3.0f} N\n"
                    f"Force:                   {target_force.whiteForce:>3.0f} N\n"
                    f"Deviation:         {deviation_symbol}\n"
                    f"====================================\n"
                )
                
                # Display monitoring circle
                for circle, label in zip(hole_stimuli, hole_labels):
                    circle.draw()
                    label.draw()
                
                # if target_lights is starting this frame...
                if target_lights.status == NOT_STARTED and t >= 0.0-frameTolerance:
                    # keep track of start time/frame for later
                    target_lights.frameNStart = frameN  # exact frame index
                    target_lights.tStart = t  # local t and not account for scr refresh
                    target_lights.tStartRefresh = tThisFlipGlobal  # on global time
                    win.timeOnFlip(target_lights, 'tStartRefresh')  # time at next scr refresh
                    # add timestamp to datafile
                    thisExp.addData('target_lights.started', t)
                    # update status
                    target_lights.status = STARTED
                    target_lights.setLights("inner", cur_target_color_rgb255)
                
                # if target_lights is active this frame...
                if target_lights.status == STARTED:
                    # update params
                    pass
                    
                
                # if target_lights is stopping this frame...
                if target_lights.status == STARTED:
                    # is it time to stop? (based on global clock, using actual start)
                    if tThisFlipGlobal > target_lights.tStartRefresh + 3.0-frameTolerance:
                        # keep track of stop time/frame for later
                        target_lights.tStop = t  # not accounting for scr refresh
                        target_lights.tStopRefresh = tThisFlipGlobal  # on global time
                        target_lights.frameNStop = frameN  # exact frame index
                        # add timestamp to datafile
                        thisExp.addData('target_lights.stopped', t)
                        # update status
                        target_lights.status = FINISHED
                        if True:
                           target_lights.turnOffLights("inner")
                
                # if target_force is starting this frame...
                if target_force.status == NOT_STARTED and t >= 0.0-frameTolerance:
                    # keep track of start time/frame for later
                    target_force.frameNStart = frameN  # exact frame index
                    target_force.tStart = t  # local t and not account for scr refresh
                    target_force.tStartRefresh = tThisFlipGlobal  # on global time
                    win.timeOnFlip(target_force, 'tStartRefresh')  # time at next scr refresh
                    # add timestamp to datafile
                    thisExp.addData('target_force.started', t)
                    # update status
                    target_force.status = STARTED
                    target_force.startForceMeasurement(50, 'both')
                
                # if target_force is active this frame...
                if target_force.status == STARTED:
                    # update params
                    pass
                    target_force.updateForceMeasurement()
                
                # if target_force is stopping this frame...
                if target_force.status == STARTED:
                    # is it time to stop? (based on global clock, using actual start)
                    if tThisFlipGlobal > target_force.tStartRefresh + 3.0-frameTolerance:
                        # keep track of stop time/frame for later
                        target_force.tStop = t  # not accounting for scr refresh
                        target_force.tStopRefresh = tThisFlipGlobal  # on global time
                        target_force.frameNStop = frameN  # exact frame index
                        # add timestamp to datafile
                        thisExp.addData('target_force.stopped', t)
                        # update status
                        target_force.status = FINISHED
                        target_force.stopForceMeasurement()
                
                # *target_display* updates
                
                # if target_display is starting this frame...
                if target_display.status == NOT_STARTED and tThisFlip >= 0.0-frameTolerance:
                    # keep track of start time/frame for later
                    target_display.frameNStart = frameN  # exact frame index
                    target_display.tStart = t  # local t and not account for scr refresh
                    target_display.tStartRefresh = tThisFlipGlobal  # on global time
                    win.timeOnFlip(target_display, 'tStartRefresh')  # time at next scr refresh
                    # add timestamp to datafile
                    thisExp.timestampOnFlip(win, 'target_display.started')
                    # update status
                    target_display.status = STARTED
                    target_display.setAutoDraw(True)
                
                # if target_display is active this frame...
                if target_display.status == STARTED:
                    # update params
                    target_display.setText(target_display_text, log=False)
                
                # if target_display is stopping this frame...
                if target_display.status == STARTED:
                    # is it time to stop? (based on global clock, using actual start)
                    if tThisFlipGlobal > target_display.tStartRefresh + 3.0-frameTolerance:
                        # keep track of stop time/frame for later
                        target_display.tStop = t  # not accounting for scr refresh
                        target_display.tStopRefresh = tThisFlipGlobal  # on global time
                        target_display.frameNStop = frameN  # exact frame index
                        # add timestamp to datafile
                        thisExp.timestampOnFlip(win, 'target_display.stopped')
                        # update status
                        target_display.status = FINISHED
                        target_display.setAutoDraw(False)
                
                # check for quit (typically the Esc key)
                if defaultKeyboard.getKeys(keyList=["escape"]):
                    thisExp.status = FINISHED
                if thisExp.status == FINISHED or endExpNow:
                    endExperiment(thisExp, win=win)
                    return
                # pause experiment here if requested
                if thisExp.status == PAUSED:
                    pauseExperiment(
                        thisExp=thisExp, 
                        win=win, 
                        timers=[routineTimer, globalClock], 
                        currentRoutine=rr_target,
                    )
                    # skip the frame we paused on
                    continue
                
                # has a Component requested the Routine to end?
                if not continueRoutine:
                    rr_target.forceEnded = routineForceEnded = True
                # has the Routine been forcibly ended?
                if rr_target.forceEnded or routineForceEnded:
                    break
                # has every Component finished?
                continueRoutine = False
                for thisComponent in rr_target.components:
                    if hasattr(thisComponent, "status") and thisComponent.status != FINISHED:
                        continueRoutine = True
                        break  # at least one component has not yet finished
                
                # refresh the screen
                if continueRoutine:  # don't flip if this routine is over or we'll get a blank screen
                    win.flip()
            
            # --- Ending Routine "rr_target" ---
            for thisComponent in rr_target.components:
                if hasattr(thisComponent, "setAutoDraw"):
                    thisComponent.setAutoDraw(False)
            # store stop times for rr_target
            rr_target.tStop = globalClock.getTime(format='float')
            rr_target.tStopRefresh = tThisFlipGlobal
            thisExp.addData('rr_target.stopped', rr_target.tStop)
            if True:
                target_lights.turnOffLights("inner")
            oo_level.addData('target_force.rate', 50)
            oo_level.addData('target_force.dynamometer', 'both')
            oo_level.addData('target_force.maxWhiteForce', target_force.maxWhiteForce)
            oo_level.addData('target_force.maxBlueForce', target_force.maxBlueForce)
            if True:
                _raw_path = thisExp.dataFileName + '_force_long.tsv'
                _write_header = not os.path.exists(_raw_path)
                _loop = oo_level
                _trial_index = _loop.thisN if _loop is not None and hasattr(_loop, 'thisN') else -1
                _trial_name = _loop.name if _loop is not None and hasattr(_loop, 'name') else ''
                _identifier = str(target_force) if target_force is not None else ''
                _records = target_force.forceRows if hasattr(target_force, 'forceRows') else []
                _times = target_force.times
                _white = target_force.whiteForceValues
                _blue = target_force.blueForceValues
                _n = max(len(_times), len(_white), len(_blue))
                with open(_raw_path, 'a', encoding='utf-8') as _f:
                    if _write_header:
                        _f.write('participant	session	routine	component	trial_index	trial_name	identifier	sample_index	white_time	blue_time	time	white_force	blue_force	white_force_raw_counts	blue_force_raw_counts\n')
                    for _i, _record in enumerate(_records):
                        _row = [
                            expInfo.get("participant", ""),
                            expInfo.get("session", ""),
                            'rr_target',
                            'target_force',
                            _trial_index,
                            _trial_name,
                            _identifier,
                            _i,
                            _record['white_time'],
                            _record['blue_time'],
                            _record['time'],
                            _record['white_force'],
                            _record['blue_force'],
                            _record['white_force_raw_counts'] if _record['white_force_raw_counts'] is not None else '',
                            _record['blue_force_raw_counts'] if _record['blue_force_raw_counts'] is not None else '',
                        ]
                        _f.write('	'.join(str(_v) for _v in _row) + '\n')
            # using non-slip timing so subtract the expected duration of this Routine (unless ended on request)
            if rr_target.maxDurationReached:
                routineTimer.addTime(-rr_target.maxDuration)
            elif rr_target.forceEnded:
                routineTimer.reset()
            else:
                routineTimer.addTime(-3.000000)
            
            # set up handler to look after randomisation of conditions etc
            oo_palette = data.TrialHandler2(
                name='oo_palette',
                nReps=NUM_PALETTES, 
                method='random', 
                extraInfo=expInfo, 
                originPath=-1, 
                trialList=[None], 
                seed=None, 
                isTrials=True, 
            )
            thisExp.addLoop(oo_palette)  # add the loop to the experiment
            thisOo_palette = oo_palette.trialList[0]  # so we can initialise stimuli with some values
            # abbreviate parameter names if possible (e.g. rgb = thisOo_palette.rgb)
            if thisOo_palette != None:
                for paramName in thisOo_palette:
                    globals()[paramName] = thisOo_palette[paramName]
            if thisSession is not None:
                # if running in a Session with a Liaison client, send data up to now
                thisSession.sendExperimentData()
            
            for thisOo_palette in oo_palette:
                oo_palette.status = STARTED
                if hasattr(thisOo_palette, 'status'):
                    thisOo_palette.status = STARTED
                currentLoop = oo_palette
                thisExp.timestampOnFlip(win, 'thisRow.t', format=globalClock.format)
                if thisSession is not None:
                    # if running in a Session with a Liaison client, send data up to now
                    thisSession.sendExperimentData()
                # abbreviate parameter names if possible (e.g. rgb = thisOo_palette.rgb)
                if thisOo_palette != None:
                    for paramName in thisOo_palette:
                        globals()[paramName] = thisOo_palette[paramName]
                
                # --- Prepare to start Routine "rr_trial" ---
                # create an object to store info about Routine rr_trial
                rr_trial = data.Routine(
                    name='rr_trial',
                    components=[trial_reed, trial_lights, trial_force, trial_display],
                )
                rr_trial.status = NOT_STARTED
                continueRoutine = True
                # update component parameters for each repeat
                # Run 'Begin Routine' code from trial_color_code
                # ===== CONDITION SETUP =====
                cur_palette             = blocks[oo_block.thisN][oo_level.thisN]['palettes'][oo_palette.thisN]
                cur_palette_has_target  = blocks[oo_block.thisN][oo_level.thisN]['palette_has_target'][oo_palette.thisN]
                cur_palette_target_hole = blocks[oo_block.thisN][oo_level.thisN]['palette_target_hole'][oo_palette.thisN]
                
                prev_reed_holes = None
                prev_reed_actions = None
                
                
                # =============================================================================
                # AUDIO - BEGIN ROUTINE
                # =============================================================================
                with _state_lock:
                    _target_amp  = 0.0
                    _target_freq = FREQ_LOW
                
                targetForce     = cur_physical_demand_level * mvc
                lower_threshold = max(targetForce - TOLERANCE, 0)
                upper_threshold = targetForce + TOLERANCE
                
                toneState    = 0
                _lastSwitchT = -1e9
                _last_beep_t = -1e9  # tracks last beep time
                
                
                # ===== DISPLAY SETUP =====
                
                trial_display.alignText = "left"
                trial_display.bold = True
                
                # Color the display circle
                for i, circle in enumerate(hole_stimuli):
                    circle.fillColor = cur_palette[i]
                # store start times for rr_trial
                rr_trial.tStartRefresh = win.getFutureFlipTime(clock=globalClock)
                rr_trial.tStart = globalClock.getTime(format='float')
                rr_trial.status = STARTED
                thisExp.addData('rr_trial.started', rr_trial.tStart)
                rr_trial.maxDuration = None
                # keep track of which components have finished
                rr_trialComponents = rr_trial.components
                for thisComponent in rr_trial.components:
                    thisComponent.tStart = None
                    thisComponent.tStop = None
                    thisComponent.tStartRefresh = None
                    thisComponent.tStopRefresh = None
                    if hasattr(thisComponent, 'status'):
                        thisComponent.status = NOT_STARTED
                # reset timers
                t = 0
                _timeToFirstFrame = win.getFutureFlipTime(clock="now")
                frameN = -1
                
                # --- Run Routine "rr_trial" ---
                thisExp.currentRoutine = rr_trial
                rr_trial.forceEnded = routineForceEnded = not continueRoutine
                while continueRoutine:
                    # if trial has changed, end Routine now
                    if hasattr(thisOo_palette, 'status') and thisOo_palette.status == STOPPING:
                        continueRoutine = False
                    # get current time
                    t = routineTimer.getTime()
                    tThisFlip = win.getFutureFlipTime(clock=routineTimer)
                    tThisFlipGlobal = win.getFutureFlipTime(clock=None)
                    frameN = frameN + 1  # number of completed frames (so 0 is the first frame)
                    # update/draw components on each frame
                    # Run 'Each Frame' code from trial_color_code
                    # =============================================================================
                    # AUDIO - EACH FRAME
                    # =============================================================================
                    if trial_force.whiteForce is not None:
                        if trial_force.whiteForce > upper_threshold:
                            proposed = 2
                        elif trial_force.whiteForce < lower_threshold:
                            proposed = 1
                        else:
                            proposed = 0
                    
                        if proposed != toneState and (t - _lastSwitchT) >= STATE_HOLD:
                            toneState    = proposed
                            _lastSwitchT = t
                            if toneState == 2:
                                freq, amp = FREQ_HIGH, AMP_MAX
                            elif toneState == 1:
                                freq, amp = FREQ_LOW, AMP_MAX
                            else:
                                freq, amp = FREQ_LOW, 0.0
                            with _state_lock:
                                _target_freq = freq
                                _target_amp  = amp
                    else:
                        print("Force is None!")
                    
                    
                    # ===== DISPLAY SETUP =====
                    if trial_force.whiteForce > upper_threshold:
                        deviation_symbol = "\u2191\u2191\u2191"  # ↑
                    elif trial_force.whiteForce < lower_threshold:
                        deviation_symbol = "\u2193\u2193\u2193"  # ↓
                    else:
                        deviation_symbol = "\u2013\u2013\u2013"  # – (on target)
                    
                    trial_display_text = (
                        f"============ TRIAL INFO ============\n"
                        f"Target Color:      {cur_target_color_name}\n"
                        f"Physical demand:   {cur_physical_demand_level}\n"
                        f"Cognitive demand:  {cur_cognitive_demand_level}\n"
                        f"------------------------------------\n"
                        f"MVC:               {mvc:>4.0f} N\n"
                        f"Target:            {lower_threshold:>3.0f} \u007C {targetForce:>3.0f} \u007C {upper_threshold:>3.0f} N\n"
                        f"Force:                   {trial_force.whiteForce:>3.0f} N\n"
                        f"Deviation:         {deviation_symbol}\n"
                        f"====================================\n"
                    )
                    
                    for circle, label in zip(hole_stimuli, hole_labels):
                        circle.draw()
                        label.draw()
                    
                    # Highlights reed contacts
                    last_selected = trial_reed.reedHoles[-1] if trial_reed.reedHoles else None
                    
                    for i, (circle, label) in enumerate(zip(hole_stimuli, hole_labels)):
                        # target highlight
                        if cur_palette_has_target and i == cur_palette_target_hole:
                            circle.lineColor = 'yellow'
                            circle.lineWidth = 3
                        # last selected highlight
                        elif i == last_selected:
                            circle.lineColor = 'white'
                            circle.lineWidth = 3
                        else:
                            circle.lineColor = None
                        circle.draw()
                        label.draw()
                    
                    # if trial_reed is starting this frame...
                    if trial_reed.status == NOT_STARTED and t >= 0.0-frameTolerance:
                        # keep track of start time/frame for later
                        trial_reed.frameNStart = frameN  # exact frame index
                        trial_reed.tStart = t  # local t and not account for scr refresh
                        trial_reed.tStartRefresh = tThisFlipGlobal  # on global time
                        win.timeOnFlip(trial_reed, 'tStartRefresh')  # time at next scr refresh
                        # add timestamp to datafile
                        thisExp.addData('trial_reed.started', t)
                        # update status
                        trial_reed.status = STARTED
                        trial_reed.startReedMeasurement(100, "inner")
                    
                    # if trial_reed is active this frame...
                    if trial_reed.status == STARTED:
                        # update params
                        pass
                        trial_reed.updateReedMeasurement()
                        if False and len(trial_reed.reedTimes) > 0:
                            trial_reed.stopReedMeasurement()
                            continueRoutine = False
                    
                    # if trial_reed is stopping this frame...
                    if trial_reed.status == STARTED:
                        # is it time to stop? (based on global clock, using actual start)
                        if tThisFlipGlobal > trial_reed.tStartRefresh + PALETTE_DURATION_SEC + 1-frameTolerance:
                            # keep track of stop time/frame for later
                            trial_reed.tStop = t  # not accounting for scr refresh
                            trial_reed.tStopRefresh = tThisFlipGlobal  # on global time
                            trial_reed.frameNStop = frameN  # exact frame index
                            # add timestamp to datafile
                            thisExp.addData('trial_reed.stopped', t)
                            # update status
                            trial_reed.status = FINISHED
                            trial_reed.stopReedMeasurement()
                    
                    # if trial_lights is starting this frame...
                    if trial_lights.status == NOT_STARTED and t >= 0.0-frameTolerance:
                        # keep track of start time/frame for later
                        trial_lights.frameNStart = frameN  # exact frame index
                        trial_lights.tStart = t  # local t and not account for scr refresh
                        trial_lights.tStartRefresh = tThisFlipGlobal  # on global time
                        win.timeOnFlip(trial_lights, 'tStartRefresh')  # time at next scr refresh
                        # add timestamp to datafile
                        thisExp.addData('trial_lights.started', t)
                        # update status
                        trial_lights.status = STARTED
                        trial_lights.setLights("inner", cur_palette)
                    
                    # if trial_lights is active this frame...
                    if trial_lights.status == STARTED:
                        # update params
                        pass
                        
                    
                    # if trial_lights is stopping this frame...
                    if trial_lights.status == STARTED:
                        # is it time to stop? (based on global clock, using actual start)
                        if tThisFlipGlobal > trial_lights.tStartRefresh + PALETTE_DURATION_SEC-frameTolerance:
                            # keep track of stop time/frame for later
                            trial_lights.tStop = t  # not accounting for scr refresh
                            trial_lights.tStopRefresh = tThisFlipGlobal  # on global time
                            trial_lights.frameNStop = frameN  # exact frame index
                            # add timestamp to datafile
                            thisExp.addData('trial_lights.stopped', t)
                            # update status
                            trial_lights.status = FINISHED
                            if True:
                               trial_lights.turnOffLights("inner")
                    
                    # if trial_force is starting this frame...
                    if trial_force.status == NOT_STARTED and t >= 0.0-frameTolerance:
                        # keep track of start time/frame for later
                        trial_force.frameNStart = frameN  # exact frame index
                        trial_force.tStart = t  # local t and not account for scr refresh
                        trial_force.tStartRefresh = tThisFlipGlobal  # on global time
                        win.timeOnFlip(trial_force, 'tStartRefresh')  # time at next scr refresh
                        # add timestamp to datafile
                        thisExp.addData('trial_force.started', t)
                        # update status
                        trial_force.status = STARTED
                        trial_force.startForceMeasurement(50, 'both')
                    
                    # if trial_force is active this frame...
                    if trial_force.status == STARTED:
                        # update params
                        pass
                        trial_force.updateForceMeasurement()
                    
                    # if trial_force is stopping this frame...
                    if trial_force.status == STARTED:
                        # is it time to stop? (based on global clock, using actual start)
                        if tThisFlipGlobal > trial_force.tStartRefresh + PALETTE_DURATION_SEC + 1-frameTolerance:
                            # keep track of stop time/frame for later
                            trial_force.tStop = t  # not accounting for scr refresh
                            trial_force.tStopRefresh = tThisFlipGlobal  # on global time
                            trial_force.frameNStop = frameN  # exact frame index
                            # add timestamp to datafile
                            thisExp.addData('trial_force.stopped', t)
                            # update status
                            trial_force.status = FINISHED
                            trial_force.stopForceMeasurement()
                    
                    # *trial_display* updates
                    
                    # if trial_display is starting this frame...
                    if trial_display.status == NOT_STARTED and tThisFlip >= 0.0-frameTolerance:
                        # keep track of start time/frame for later
                        trial_display.frameNStart = frameN  # exact frame index
                        trial_display.tStart = t  # local t and not account for scr refresh
                        trial_display.tStartRefresh = tThisFlipGlobal  # on global time
                        win.timeOnFlip(trial_display, 'tStartRefresh')  # time at next scr refresh
                        # add timestamp to datafile
                        thisExp.timestampOnFlip(win, 'trial_display.started')
                        # update status
                        trial_display.status = STARTED
                        trial_display.setAutoDraw(True)
                    
                    # if trial_display is active this frame...
                    if trial_display.status == STARTED:
                        # update params
                        trial_display.setText(trial_display_text, log=False)
                    
                    # if trial_display is stopping this frame...
                    if trial_display.status == STARTED:
                        # is it time to stop? (based on global clock, using actual start)
                        if tThisFlipGlobal > trial_display.tStartRefresh + PALETTE_DURATION_SEC + 1-frameTolerance:
                            # keep track of stop time/frame for later
                            trial_display.tStop = t  # not accounting for scr refresh
                            trial_display.tStopRefresh = tThisFlipGlobal  # on global time
                            trial_display.frameNStop = frameN  # exact frame index
                            # add timestamp to datafile
                            thisExp.timestampOnFlip(win, 'trial_display.stopped')
                            # update status
                            trial_display.status = FINISHED
                            trial_display.setAutoDraw(False)
                    
                    # check for quit (typically the Esc key)
                    if defaultKeyboard.getKeys(keyList=["escape"]):
                        thisExp.status = FINISHED
                    if thisExp.status == FINISHED or endExpNow:
                        endExperiment(thisExp, win=win)
                        return
                    # pause experiment here if requested
                    if thisExp.status == PAUSED:
                        pauseExperiment(
                            thisExp=thisExp, 
                            win=win, 
                            timers=[routineTimer, globalClock], 
                            currentRoutine=rr_trial,
                        )
                        # skip the frame we paused on
                        continue
                    
                    # has a Component requested the Routine to end?
                    if not continueRoutine:
                        rr_trial.forceEnded = routineForceEnded = True
                    # has the Routine been forcibly ended?
                    if rr_trial.forceEnded or routineForceEnded:
                        break
                    # has every Component finished?
                    continueRoutine = False
                    for thisComponent in rr_trial.components:
                        if hasattr(thisComponent, "status") and thisComponent.status != FINISHED:
                            continueRoutine = True
                            break  # at least one component has not yet finished
                    
                    # refresh the screen
                    if continueRoutine:  # don't flip if this routine is over or we'll get a blank screen
                        win.flip()
                
                # --- Ending Routine "rr_trial" ---
                for thisComponent in rr_trial.components:
                    if hasattr(thisComponent, "setAutoDraw"):
                        thisComponent.setAutoDraw(False)
                # store stop times for rr_trial
                rr_trial.tStop = globalClock.getTime(format='float')
                rr_trial.tStopRefresh = tThisFlipGlobal
                thisExp.addData('rr_trial.stopped', rr_trial.tStop)
                # Run 'End Routine' code from trial_color_code
                # Determine d-prime metrics
                is_hit               = int(cur_palette_has_target and cur_palette_target_hole in trial_reed.reedHoles)
                is_false_alarm       = int(not cur_palette_has_target and len(trial_reed.reedHoles) > 0)
                is_miss              = int(cur_palette_has_target and cur_palette_target_hole not in trial_reed.reedHoles)
                is_correct_rejection = int(not cur_palette_has_target and len(trial_reed.reedHoles) == 0)
                
                # Accumulate d-prime metrics
                hits               += is_hit
                false_alarms       += is_false_alarm
                misses             += is_miss
                correct_rejections += is_correct_rejection
                
                # Get the first insertion event (action == 1) across all reed events
                first_insert_idx = next((h for h, a in zip(trial_reed.reedHoles, trial_reed.reedActions) if a == 1), None)
                first_insert_rt  = next((t for t, a in zip(trial_reed.reedTimes,  trial_reed.reedActions) if a == 1), None)
                
                # Write hole-level data: mark which hole was first selected and its RT
                for hole_idx in range(NUM_DISTRACTORS):
                    i = row_index[(oo_block.thisN, oo_level.thisN, oo_palette.thisN, hole_idx)]
                    rows[i]["is_selected"] = int(hole_idx == first_insert_idx)
                    rows[i]["selected_rt"] = first_insert_rt if hole_idx == first_insert_idx else None
                
                first_insert_rt = next((t for t, a in zip(trial_reed.reedTimes, trial_reed.reedActions) if a == 1), None)
                rows[i]["selected_rt"] = first_insert_rt
                
                # Write palette-level data
                for i in palette_row_index[(oo_block.thisN, oo_level.thisN, oo_palette.thisN)]:
                    rows[i]["is_hit"]               = is_hit
                    rows[i]["is_false_alarm"]       = is_false_alarm
                    rows[i]["is_miss"]              = is_miss
                    rows[i]["is_correct_rejection"] = is_correct_rejection
                
                #Daten anhängen
                # paletteResponseList = sorted(paletteResponses)
                # currentLoop.addData('paletteResponseCount', len(paletteResponseList))
                # currentLoop.addData('paletteResponses', paletteResponseList)
                # currentLoop.addData('paletteResponseEvents', paletteResponseEvents)
                # currentLoop.addData('hit', hit)
                # currentLoop.addData('falsealarm', fa)
                # currentLoop.addData('correct rejection', cr)
                # currentLoop.addData('miss', miss)
                
                trial_reed.stopReedMeasurement()
                oo_palette.addData('trial_reed.rate', 100)
                oo_palette.addData('trial_reed.holes', "inner")
                oo_palette.addData('trial_reed.reedTimes', trial_reed.reedTimes)
                oo_palette.addData('trial_reed.reedHoles', trial_reed.reedHoles)
                oo_palette.addData('trial_reed.reedActions', trial_reed.reedActions)
                oo_palette.addData('trial_reed.reedSummary', trial_reed.reedSummary)
                oo_palette.addData('trial_reed.reedCurrentStates', trial_reed.reedCurrentStates)
                oo_palette.addData('trial_reed.reedActiveHoles', trial_reed.reedActiveHoles)
                oo_palette.addData('trial_reed.reedNewInsertions', trial_reed.reedNewInsertions)
                oo_palette.addData('trial_reed.reedNewRemovals', trial_reed.reedNewRemovals)
                oo_palette.addData('trial_reed.reedLatestEvent', trial_reed.reedLatestEvent)
                oo_palette.addData('trial_reed.reedFrameTimes', trial_reed.reedFrameTimes)
                oo_palette.addData('trial_reed.reedFrameStates', trial_reed.reedFrameStates)
                oo_palette.addData('trial_reed.reedFrameActiveHoles', trial_reed.reedFrameActiveHoles)
                if True:
                    trial_lights.turnOffLights("inner")
                oo_palette.addData('trial_force.rate', 50)
                oo_palette.addData('trial_force.dynamometer', 'both')
                oo_palette.addData('trial_force.maxWhiteForce', trial_force.maxWhiteForce)
                oo_palette.addData('trial_force.maxBlueForce', trial_force.maxBlueForce)
                if True:
                    _raw_path = thisExp.dataFileName + '_force_long.tsv'
                    _write_header = not os.path.exists(_raw_path)
                    _loop = oo_palette
                    _trial_index = _loop.thisN if _loop is not None and hasattr(_loop, 'thisN') else -1
                    _trial_name = _loop.name if _loop is not None and hasattr(_loop, 'name') else ''
                    _identifier = str(trial_force) if trial_force is not None else ''
                    _records = trial_force.forceRows if hasattr(trial_force, 'forceRows') else []
                    _times = trial_force.times
                    _white = trial_force.whiteForceValues
                    _blue = trial_force.blueForceValues
                    _n = max(len(_times), len(_white), len(_blue))
                    with open(_raw_path, 'a', encoding='utf-8') as _f:
                        if _write_header:
                            _f.write('participant	session	routine	component	trial_index	trial_name	identifier	sample_index	white_time	blue_time	time	white_force	blue_force	white_force_raw_counts	blue_force_raw_counts\n')
                        for _i, _record in enumerate(_records):
                            _row = [
                                expInfo.get("participant", ""),
                                expInfo.get("session", ""),
                                'rr_trial',
                                'trial_force',
                                _trial_index,
                                _trial_name,
                                _identifier,
                                _i,
                                _record['white_time'],
                                _record['blue_time'],
                                _record['time'],
                                _record['white_force'],
                                _record['blue_force'],
                                _record['white_force_raw_counts'] if _record['white_force_raw_counts'] is not None else '',
                                _record['blue_force_raw_counts'] if _record['blue_force_raw_counts'] is not None else '',
                            ]
                            _f.write('	'.join(str(_v) for _v in _row) + '\n')
                # the Routine "rr_trial" was not non-slip safe, so reset the non-slip timer
                routineTimer.reset()
                # mark thisOo_palette as finished
                if hasattr(thisOo_palette, 'status'):
                    thisOo_palette.status = FINISHED
                # if awaiting a pause, pause now
                if oo_palette.status == PAUSED:
                    thisExp.status = PAUSED
                    pauseExperiment(
                        thisExp=thisExp, 
                        win=win, 
                        timers=[globalClock], 
                    )
                    # once done pausing, restore running status
                    oo_palette.status = STARTED
                thisExp.nextEntry()
                
            # completed NUM_PALETTES repeats of 'oo_palette'
            oo_palette.status = FINISHED
            
            if thisSession is not None:
                # if running in a Session with a Liaison client, send data up to now
                thisSession.sendExperimentData()
            
            # --- Prepare to start Routine "rr_dprime" ---
            # create an object to store info about Routine rr_dprime
            rr_dprime = data.Routine(
                name='rr_dprime',
                components=[feedback_keyboard, feedback_display],
            )
            rr_dprime.status = NOT_STARTED
            continueRoutine = True
            # update component parameters for each repeat
            # Run 'Begin Routine' code from feedback_code
            #dprime Calculation
            H = hits + misses
            F = false_alarms + correct_rejections
            
            hit_rate = (hits+0.5)/(H+1) if H>0 else 0.5
            fa_rate = (false_alarms+0.5)/(F+1) if F>0 else 0.5
            d_prime = norm.ppf(hit_rate) - norm.ppf(fa_rate)
            
            currentLoop.addData('d_prime', d_prime)
            
            # Write trial-level data
            for palette_idx in range(NUM_PALETTES):
                for i in palette_row_index[(oo_block.thisN, oo_level.thisN, palette_idx)]:
                    rows[i]["hits"]               = hits
                    rows[i]["false_alarms"]       = false_alarms
                    rows[i]["misses"]             = misses
                    rows[i]["correct_rejections"] = correct_rejections
                    rows[i]["d_prime"]            = d_prime
            
            # ===== DISPLAY SETUP =====
            
            feedback_display.alignText = "left"
            feedback_display.bold = True
            
            feedback_display_text = (
                f"=========== FEEDBACK ===========\n"
                f"d-prime:           {d_prime:.2f}\n"
                f"Hits:              {hits}\n"
                f"Misses:            {misses}\n"
                f"False Alarms:      {false_alarms}\n"
                f"Correct Rej.:      {correct_rejections}\n"
                f"================================\n"
            )
            # create starting attributes for feedback_keyboard
            feedback_keyboard.keys = []
            feedback_keyboard.rt = []
            _feedback_keyboard_allKeys = []
            feedback_display.setText(feedback_display_text)
            # store start times for rr_dprime
            rr_dprime.tStartRefresh = win.getFutureFlipTime(clock=globalClock)
            rr_dprime.tStart = globalClock.getTime(format='float')
            rr_dprime.status = STARTED
            thisExp.addData('rr_dprime.started', rr_dprime.tStart)
            rr_dprime.maxDuration = None
            # keep track of which components have finished
            rr_dprimeComponents = rr_dprime.components
            for thisComponent in rr_dprime.components:
                thisComponent.tStart = None
                thisComponent.tStop = None
                thisComponent.tStartRefresh = None
                thisComponent.tStopRefresh = None
                if hasattr(thisComponent, 'status'):
                    thisComponent.status = NOT_STARTED
            # reset timers
            t = 0
            _timeToFirstFrame = win.getFutureFlipTime(clock="now")
            frameN = -1
            
            # --- Run Routine "rr_dprime" ---
            thisExp.currentRoutine = rr_dprime
            rr_dprime.forceEnded = routineForceEnded = not continueRoutine
            while continueRoutine:
                # if trial has changed, end Routine now
                if hasattr(thisOo_level, 'status') and thisOo_level.status == STOPPING:
                    continueRoutine = False
                # get current time
                t = routineTimer.getTime()
                tThisFlip = win.getFutureFlipTime(clock=routineTimer)
                tThisFlipGlobal = win.getFutureFlipTime(clock=None)
                frameN = frameN + 1  # number of completed frames (so 0 is the first frame)
                # update/draw components on each frame
                
                # *feedback_keyboard* updates
                waitOnFlip = False
                
                # if feedback_keyboard is starting this frame...
                if feedback_keyboard.status == NOT_STARTED and tThisFlip >= 0.0-frameTolerance:
                    # keep track of start time/frame for later
                    feedback_keyboard.frameNStart = frameN  # exact frame index
                    feedback_keyboard.tStart = t  # local t and not account for scr refresh
                    feedback_keyboard.tStartRefresh = tThisFlipGlobal  # on global time
                    win.timeOnFlip(feedback_keyboard, 'tStartRefresh')  # time at next scr refresh
                    # add timestamp to datafile
                    thisExp.timestampOnFlip(win, 'feedback_keyboard.started')
                    # update status
                    feedback_keyboard.status = STARTED
                    # keyboard checking is just starting
                    waitOnFlip = True
                    win.callOnFlip(feedback_keyboard.clock.reset)  # t=0 on next screen flip
                    win.callOnFlip(feedback_keyboard.clearEvents, eventType='keyboard')  # clear events on next screen flip
                if feedback_keyboard.status == STARTED and not waitOnFlip:
                    theseKeys = feedback_keyboard.getKeys(keyList=['space'], ignoreKeys=["escape"], waitRelease=False)
                    _feedback_keyboard_allKeys.extend(theseKeys)
                    if len(_feedback_keyboard_allKeys):
                        feedback_keyboard.keys = _feedback_keyboard_allKeys[-1].name  # just the last key pressed
                        feedback_keyboard.rt = _feedback_keyboard_allKeys[-1].rt
                        feedback_keyboard.duration = _feedback_keyboard_allKeys[-1].duration
                        # a response ends the routine
                        continueRoutine = False
                
                # *feedback_display* updates
                
                # if feedback_display is starting this frame...
                if feedback_display.status == NOT_STARTED and tThisFlip >= 0-frameTolerance:
                    # keep track of start time/frame for later
                    feedback_display.frameNStart = frameN  # exact frame index
                    feedback_display.tStart = t  # local t and not account for scr refresh
                    feedback_display.tStartRefresh = tThisFlipGlobal  # on global time
                    win.timeOnFlip(feedback_display, 'tStartRefresh')  # time at next scr refresh
                    # add timestamp to datafile
                    thisExp.timestampOnFlip(win, 'feedback_display.started')
                    # update status
                    feedback_display.status = STARTED
                    feedback_display.setAutoDraw(True)
                
                # if feedback_display is active this frame...
                if feedback_display.status == STARTED:
                    # update params
                    pass
                
                # check for quit (typically the Esc key)
                if defaultKeyboard.getKeys(keyList=["escape"]):
                    thisExp.status = FINISHED
                if thisExp.status == FINISHED or endExpNow:
                    endExperiment(thisExp, win=win)
                    return
                # pause experiment here if requested
                if thisExp.status == PAUSED:
                    pauseExperiment(
                        thisExp=thisExp, 
                        win=win, 
                        timers=[routineTimer, globalClock], 
                        currentRoutine=rr_dprime,
                    )
                    # skip the frame we paused on
                    continue
                
                # has a Component requested the Routine to end?
                if not continueRoutine:
                    rr_dprime.forceEnded = routineForceEnded = True
                # has the Routine been forcibly ended?
                if rr_dprime.forceEnded or routineForceEnded:
                    break
                # has every Component finished?
                continueRoutine = False
                for thisComponent in rr_dprime.components:
                    if hasattr(thisComponent, "status") and thisComponent.status != FINISHED:
                        continueRoutine = True
                        break  # at least one component has not yet finished
                
                # refresh the screen
                if continueRoutine:  # don't flip if this routine is over or we'll get a blank screen
                    win.flip()
            
            # --- Ending Routine "rr_dprime" ---
            for thisComponent in rr_dprime.components:
                if hasattr(thisComponent, "setAutoDraw"):
                    thisComponent.setAutoDraw(False)
            # store stop times for rr_dprime
            rr_dprime.tStop = globalClock.getTime(format='float')
            rr_dprime.tStopRefresh = tThisFlipGlobal
            thisExp.addData('rr_dprime.stopped', rr_dprime.tStop)
            # check responses
            if feedback_keyboard.keys in ['', [], None]:  # No response was made
                feedback_keyboard.keys = None
            oo_level.addData('feedback_keyboard.keys',feedback_keyboard.keys)
            if feedback_keyboard.keys != None:  # we had a response
                oo_level.addData('feedback_keyboard.rt', feedback_keyboard.rt)
                oo_level.addData('feedback_keyboard.duration', feedback_keyboard.duration)
            # the Routine "rr_dprime" was not non-slip safe, so reset the non-slip timer
            routineTimer.reset()
            
            # set up handler to look after randomisation of conditions etc
            oo_slider = data.TrialHandler2(
                name='oo_slider',
                nReps=1, 
                method='sequential', 
                extraInfo=expInfo, 
                originPath=-1, 
                trialList=data.importConditions('oo_slider.xlsx'), 
                seed=None, 
                isTrials=True, 
            )
            thisExp.addLoop(oo_slider)  # add the loop to the experiment
            thisOo_slider = oo_slider.trialList[0]  # so we can initialise stimuli with some values
            # abbreviate parameter names if possible (e.g. rgb = thisOo_slider.rgb)
            if thisOo_slider != None:
                for paramName in thisOo_slider:
                    globals()[paramName] = thisOo_slider[paramName]
            if thisSession is not None:
                # if running in a Session with a Liaison client, send data up to now
                thisSession.sendExperimentData()
            
            for thisOo_slider in oo_slider:
                oo_slider.status = STARTED
                if hasattr(thisOo_slider, 'status'):
                    thisOo_slider.status = STARTED
                currentLoop = oo_slider
                thisExp.timestampOnFlip(win, 'thisRow.t', format=globalClock.format)
                if thisSession is not None:
                    # if running in a Session with a Liaison client, send data up to now
                    thisSession.sendExperimentData()
                # abbreviate parameter names if possible (e.g. rgb = thisOo_slider.rgb)
                if thisOo_slider != None:
                    for paramName in thisOo_slider:
                        globals()[paramName] = thisOo_slider[paramName]
                
                # --- Prepare to start Routine "rr_rating" ---
                # create an object to store info about Routine rr_rating
                rr_rating = data.Routine(
                    name='rr_rating',
                    components=[rating_slider, rating_keyboard, rating_display],
                )
                rr_rating.status = NOT_STARTED
                continueRoutine = True
                # update component parameters for each repeat
                rating_slider.reset()
                # create starting attributes for rating_keyboard
                rating_keyboard.keys = []
                rating_keyboard.rt = []
                _rating_keyboard_allKeys = []
                rating_display.setText(cond_slider)
                # store start times for rr_rating
                rr_rating.tStartRefresh = win.getFutureFlipTime(clock=globalClock)
                rr_rating.tStart = globalClock.getTime(format='float')
                rr_rating.status = STARTED
                thisExp.addData('rr_rating.started', rr_rating.tStart)
                rr_rating.maxDuration = None
                # keep track of which components have finished
                rr_ratingComponents = rr_rating.components
                for thisComponent in rr_rating.components:
                    thisComponent.tStart = None
                    thisComponent.tStop = None
                    thisComponent.tStartRefresh = None
                    thisComponent.tStopRefresh = None
                    if hasattr(thisComponent, 'status'):
                        thisComponent.status = NOT_STARTED
                # reset timers
                t = 0
                _timeToFirstFrame = win.getFutureFlipTime(clock="now")
                frameN = -1
                
                # --- Run Routine "rr_rating" ---
                thisExp.currentRoutine = rr_rating
                rr_rating.forceEnded = routineForceEnded = not continueRoutine
                while continueRoutine:
                    # if trial has changed, end Routine now
                    if hasattr(thisOo_slider, 'status') and thisOo_slider.status == STOPPING:
                        continueRoutine = False
                    # get current time
                    t = routineTimer.getTime()
                    tThisFlip = win.getFutureFlipTime(clock=routineTimer)
                    tThisFlipGlobal = win.getFutureFlipTime(clock=None)
                    frameN = frameN + 1  # number of completed frames (so 0 is the first frame)
                    # update/draw components on each frame
                    
                    # *rating_slider* updates
                    
                    # if rating_slider is starting this frame...
                    if rating_slider.status == NOT_STARTED and tThisFlip >= 0.0-frameTolerance:
                        # keep track of start time/frame for later
                        rating_slider.frameNStart = frameN  # exact frame index
                        rating_slider.tStart = t  # local t and not account for scr refresh
                        rating_slider.tStartRefresh = tThisFlipGlobal  # on global time
                        win.timeOnFlip(rating_slider, 'tStartRefresh')  # time at next scr refresh
                        # add timestamp to datafile
                        thisExp.timestampOnFlip(win, 'rating_slider.started')
                        # update status
                        rating_slider.status = STARTED
                        rating_slider.setAutoDraw(True)
                    
                    # if rating_slider is active this frame...
                    if rating_slider.status == STARTED:
                        # update params
                        pass
                    
                    # *rating_keyboard* updates
                    waitOnFlip = False
                    
                    # if rating_keyboard is starting this frame...
                    if rating_keyboard.status == NOT_STARTED and tThisFlip >= 0.0-frameTolerance:
                        # keep track of start time/frame for later
                        rating_keyboard.frameNStart = frameN  # exact frame index
                        rating_keyboard.tStart = t  # local t and not account for scr refresh
                        rating_keyboard.tStartRefresh = tThisFlipGlobal  # on global time
                        win.timeOnFlip(rating_keyboard, 'tStartRefresh')  # time at next scr refresh
                        # add timestamp to datafile
                        thisExp.timestampOnFlip(win, 'rating_keyboard.started')
                        # update status
                        rating_keyboard.status = STARTED
                        # keyboard checking is just starting
                        waitOnFlip = True
                        win.callOnFlip(rating_keyboard.clock.reset)  # t=0 on next screen flip
                        win.callOnFlip(rating_keyboard.clearEvents, eventType='keyboard')  # clear events on next screen flip
                    if rating_keyboard.status == STARTED and not waitOnFlip:
                        theseKeys = rating_keyboard.getKeys(keyList=['space'], ignoreKeys=["escape"], waitRelease=False)
                        _rating_keyboard_allKeys.extend(theseKeys)
                        if len(_rating_keyboard_allKeys):
                            rating_keyboard.keys = _rating_keyboard_allKeys[-1].name  # just the last key pressed
                            rating_keyboard.rt = _rating_keyboard_allKeys[-1].rt
                            rating_keyboard.duration = _rating_keyboard_allKeys[-1].duration
                            # a response ends the routine
                            continueRoutine = False
                    
                    # *rating_display* updates
                    
                    # if rating_display is starting this frame...
                    if rating_display.status == NOT_STARTED and tThisFlip >= 0.0-frameTolerance:
                        # keep track of start time/frame for later
                        rating_display.frameNStart = frameN  # exact frame index
                        rating_display.tStart = t  # local t and not account for scr refresh
                        rating_display.tStartRefresh = tThisFlipGlobal  # on global time
                        win.timeOnFlip(rating_display, 'tStartRefresh')  # time at next scr refresh
                        # add timestamp to datafile
                        thisExp.timestampOnFlip(win, 'rating_display.started')
                        # update status
                        rating_display.status = STARTED
                        rating_display.setAutoDraw(True)
                    
                    # if rating_display is active this frame...
                    if rating_display.status == STARTED:
                        # update params
                        pass
                    
                    # check for quit (typically the Esc key)
                    if defaultKeyboard.getKeys(keyList=["escape"]):
                        thisExp.status = FINISHED
                    if thisExp.status == FINISHED or endExpNow:
                        endExperiment(thisExp, win=win)
                        return
                    # pause experiment here if requested
                    if thisExp.status == PAUSED:
                        pauseExperiment(
                            thisExp=thisExp, 
                            win=win, 
                            timers=[routineTimer, globalClock], 
                            currentRoutine=rr_rating,
                        )
                        # skip the frame we paused on
                        continue
                    
                    # has a Component requested the Routine to end?
                    if not continueRoutine:
                        rr_rating.forceEnded = routineForceEnded = True
                    # has the Routine been forcibly ended?
                    if rr_rating.forceEnded or routineForceEnded:
                        break
                    # has every Component finished?
                    continueRoutine = False
                    for thisComponent in rr_rating.components:
                        if hasattr(thisComponent, "status") and thisComponent.status != FINISHED:
                            continueRoutine = True
                            break  # at least one component has not yet finished
                    
                    # refresh the screen
                    if continueRoutine:  # don't flip if this routine is over or we'll get a blank screen
                        win.flip()
                
                # --- Ending Routine "rr_rating" ---
                for thisComponent in rr_rating.components:
                    if hasattr(thisComponent, "setAutoDraw"):
                        thisComponent.setAutoDraw(False)
                # store stop times for rr_rating
                rr_rating.tStop = globalClock.getTime(format='float')
                rr_rating.tStopRefresh = tThisFlipGlobal
                thisExp.addData('rr_rating.stopped', rr_rating.tStop)
                oo_slider.addData('rating_slider.response', rating_slider.getRating())
                oo_slider.addData('rating_slider.rt', rating_slider.getRT())
                # check responses
                if rating_keyboard.keys in ['', [], None]:  # No response was made
                    rating_keyboard.keys = None
                oo_slider.addData('rating_keyboard.keys',rating_keyboard.keys)
                if rating_keyboard.keys != None:  # we had a response
                    oo_slider.addData('rating_keyboard.rt', rating_keyboard.rt)
                    oo_slider.addData('rating_keyboard.duration', rating_keyboard.duration)
                # Run 'End Routine' code from rating_code
                # Write trial-level data
                for palette_idx in range(NUM_PALETTES):
                    for i in palette_row_index[(oo_block.thisN, oo_level.thisN, palette_idx)]:
                        if cond_slider == "cognitive":
                            rows[i]["rpe_cognitive"] = rating_slider.getRating()
                        elif cond_slider == "physical":
                            rows[i]["rpe_physical"] = rating_slider.getRating()
                # the Routine "rr_rating" was not non-slip safe, so reset the non-slip timer
                routineTimer.reset()
                # mark thisOo_slider as finished
                if hasattr(thisOo_slider, 'status'):
                    thisOo_slider.status = FINISHED
                # if awaiting a pause, pause now
                if oo_slider.status == PAUSED:
                    thisExp.status = PAUSED
                    pauseExperiment(
                        thisExp=thisExp, 
                        win=win, 
                        timers=[globalClock], 
                    )
                    # once done pausing, restore running status
                    oo_slider.status = STARTED
                thisExp.nextEntry()
                
            # completed 1 repeats of 'oo_slider'
            oo_slider.status = FINISHED
            
            if thisSession is not None:
                # if running in a Session with a Liaison client, send data up to now
                thisSession.sendExperimentData()
            # mark thisOo_level as finished
            if hasattr(thisOo_level, 'status'):
                thisOo_level.status = FINISHED
            # if awaiting a pause, pause now
            if oo_level.status == PAUSED:
                thisExp.status = PAUSED
                pauseExperiment(
                    thisExp=thisExp, 
                    win=win, 
                    timers=[globalClock], 
                )
                # once done pausing, restore running status
                oo_level.status = STARTED
            thisExp.nextEntry()
            
        # completed NUM_LEVELS repeats of 'oo_level'
        oo_level.status = FINISHED
        
        if thisSession is not None:
            # if running in a Session with a Liaison client, send data up to now
            thisSession.sendExperimentData()
        # mark thisOo_block as finished
        if hasattr(thisOo_block, 'status'):
            thisOo_block.status = FINISHED
        # if awaiting a pause, pause now
        if oo_block.status == PAUSED:
            thisExp.status = PAUSED
            pauseExperiment(
                thisExp=thisExp, 
                win=win, 
                timers=[globalClock], 
            )
            # once done pausing, restore running status
            oo_block.status = STARTED
        thisExp.nextEntry()
        
    # completed NUM_BLOCKS repeats of 'oo_block'
    oo_block.status = FINISHED
    
    if thisSession is not None:
        # if running in a Session with a Liaison client, send data up to now
        thisSession.sendExperimentData()
    
    # mark experiment as finished
    endExperiment(thisExp, win=win)


def saveData(thisExp):
    """
    Save data from this experiment
    
    Parameters
    ==========
    thisExp : psychopy.data.ExperimentHandler
        Handler object for this experiment, contains the data to save and information about 
        where to save it to.
    """
    filename = thisExp.dataFileName
    # these shouldn't be strictly necessary (should auto-save)
    thisExp.saveAsWideText(filename + '.csv', delim='auto')
    thisExp.saveAsPickle(filename)


def endExperiment(thisExp, win=None):
    """
    End this experiment, performing final shut down operations.
    
    This function does NOT close the window or end the Python process - use `quit` for this.
    
    Parameters
    ==========
    thisExp : psychopy.data.ExperimentHandler
        Handler object for this experiment, contains the data to save and information about 
        where to save it to.
    win : psychopy.visual.Window
        Window for this experiment.
    """
    # stop any playback components
    if thisExp.currentRoutine is not None:
        for comp in thisExp.currentRoutine.getPlaybackComponents():
            comp.stop()
    if win is not None:
        # remove autodraw from all current components
        win.clearAutoDraw()
        # Flip one final time so any remaining win.callOnFlip() 
        # and win.timeOnFlip() tasks get executed
        win.flip()
    # return console logger level to WARNING
    logging.console.setLevel(logging.WARNING)
    # mark experiment handler as finished
    thisExp.status = FINISHED
    # run any 'at exit' functions
    for fcn in runAtExit:
        fcn()
    logging.flush()


def quit(thisExp, win=None, thisSession=None):
    """
    Fully quit, closing the window and ending the Python process.
    
    Parameters
    ==========
    win : psychopy.visual.Window
        Window to close.
    thisSession : psychopy.session.Session or None
        Handle of the Session object this experiment is being run from, if any.
    """
    thisExp.abort()  # or data files will save again on exit
    # make sure everything is closed down
    if win is not None:
        # Flip one final time so any remaining win.callOnFlip() 
        # and win.timeOnFlip() tasks get executed before quitting
        win.flip()
        win.close()
    logging.flush()
    if thisSession is not None:
        thisSession.stop()
    # terminate Python process
    core.quit()


# if running this experiment as a script...
if __name__ == '__main__':
    # call all functions in order
    expInfo = showExpInfoDlg(expInfo=expInfo)
    thisExp = setupData(expInfo=expInfo)
    logFile = setupLogging(filename=thisExp.dataFileName)
    win = setupWindow(expInfo=expInfo)
    setupDevices(expInfo=expInfo, thisExp=thisExp, win=win)
    run(
        expInfo=expInfo, 
        thisExp=thisExp, 
        win=win,
        globalClock='float'
    )
    saveData(thisExp=thisExp)
    quit(thisExp=thisExp, win=win)
